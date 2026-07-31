import os
import io
import json
from functools import wraps
from datetime import datetime, timedelta, timezone, date
import pytz

from flask import (
    Flask, render_template, redirect, url_for, request, flash, jsonify,
    abort, send_file, session
)
from flask_login import (
    login_user, logout_user, login_required, current_user
)
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename
from sqlalchemy.exc import IntegrityError
from PIL import Image
import openpyxl
from openpyxl.utils import get_column_letter
import re

from extensions import db, login_manager, csrf, limiter
from models import (
    User, Kategori, Aset, Tiket, TiketAset,
    LogStatus, HistoriAset, AktivitasLog, Maintenance,
    Peminjaman, PeminjamanAset, PeminjamanEvidence,
    Area, Gedung, Lantai, Ruangan
)
from roles import ROLE_ADMIN

# ---------------------------------------------------------------------------
# Konfigurasi Aplikasi
# ---------------------------------------------------------------------------
BASE_DIR = os.path.abspath(os.path.dirname(__file__))
UPLOAD_FOLDER = os.path.join(BASE_DIR, "static", "uploads")
ALLOWED_EXT = {"png", "jpg", "jpeg", "gif", "webp", "jpe", "jfif", "bmp", "tiff", "tif"}

# Daftar Jenis Transaksi Peminjaman, disamakan dengan nilai yang ada di
# sheet "BA Transfer" pada file Excel sumber, supaya data hasil import lama
# dan data yang diinput manual lewat website konsisten.
JENIS_TRANSAKSI_OPTIONS = [
    "Peminjaman",
    "Pengembalian",
    "Pelimpahan",
    "Pelimpahan IN",
    "Pelimpahan Out",
    "Pemindahan",
    "Penyerahan",
    "Serah Terima",
    "Surat Jalan",
]

app = Flask(__name__)

app.config["SECRET_KEY"] = os.environ.get(
    "SECRET_KEY", "dev-only-jangan-dipakai-di-production"
)

DB_USER = os.environ.get("DB_USER", "root")
DB_PASSWORD = os.environ.get("DB_PASSWORD", "")
DB_HOST = os.environ.get("DB_HOST", "localhost")
DB_NAME = os.environ.get("DB_NAME", "db_manajemen_aset")
app.config["SQLALCHEMY_DATABASE_URI"] = (
    f"mysql+pymysql://{DB_USER}:{DB_PASSWORD}@{DB_HOST}/{DB_NAME}"
)
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
app.config["MAX_CONTENT_LENGTH"] = 10 * 1024 * 1024

db.init_app(app)
login_manager.init_app(app)
csrf.init_app(app)
limiter.init_app(app)


@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))

# Set timezone ke WIB (UTC+7)
WIB = pytz.timezone('Asia/Jakarta')


# ---------------------------------------------------------------------------
# Helper
# ---------------------------------------------------------------------------
def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXT


def is_valid_image(file_storage):
    try:
        pos = file_storage.stream.tell()
        Image.open(file_storage.stream).verify()
        file_storage.stream.seek(pos)
        return True
    except Exception:
        return False


def save_upload(file_storage, prefix=""):
    if not file_storage or file_storage.filename == "":
        return None, None
    if not allowed_file(file_storage.filename):
        return None, f"Ekstensi file tidak diizinkan. Gunakan: {', '.join(ALLOWED_EXT)}"
    if not is_valid_image(file_storage):
        return None, "File yang diupload bukan gambar yang valid."
    filename = secure_filename(file_storage.filename)
    if not filename:
        return None, "Nama file tidak valid."
    unique_name = f"{prefix}{datetime.now(WIB).strftime('%Y%m%d%H%M%S%f')}_{filename}"
    filepath = os.path.join(app.config["UPLOAD_FOLDER"], unique_name)
    try:
        file_storage.save(filepath)
        return unique_name, None
    except Exception as e:
        return None, f"Gagal menyimpan file: {str(e)}"


ALLOWED_EXT_DOKUMEN = ALLOWED_EXT | {"pdf", "doc", "docx"}


def allowed_dokumen(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXT_DOKUMEN


def save_dokumen(file_storage, prefix=""):
    """Simpan file evidence/BA (Berita Acara). Boleh berupa gambar ATAU dokumen
    (pdf/doc/docx) -- berbeda dari save_upload() yang khusus gambar saja,
    karena Evidence Lampiran peminjaman aset biasanya berupa PDF hasil scan BA."""
    if not file_storage or file_storage.filename == "":
        return None, None
    if not allowed_dokumen(file_storage.filename):
        return None, f"Ekstensi file tidak diizinkan. Gunakan: {', '.join(sorted(ALLOWED_EXT_DOKUMEN))}"
    ext = file_storage.filename.rsplit(".", 1)[1].lower()
    # Kalau file gambar, tetap divalidasi isinya seperti save_upload biasa
    if ext in ALLOWED_EXT and not is_valid_image(file_storage):
        return None, "File yang diupload bukan gambar yang valid."
    filename = secure_filename(file_storage.filename)
    if not filename:
        return None, "Nama file tidak valid."
    unique_name = f"{prefix}{datetime.now(WIB).strftime('%Y%m%d%H%M%S%f')}_{filename}"
    filepath = os.path.join(app.config["UPLOAD_FOLDER"], unique_name)
    try:
        file_storage.save(filepath)
        return unique_name, None
    except Exception as e:
        return None, f"Gagal menyimpan file: {str(e)}"


def save_evidence_pdf(file_storage):
    """Simpan evidence Peminjaman -- WAJIB berupa file PDF (.pdf).

    Dipakai di seluruh alur evidence Peminjaman: tambah data peminjaman,
    tambah evidence tambahan, dan konfirmasi perpanjangan. Berbeda dari
    save_dokumen() (yang masih membolehkan gambar/doc/docx), fungsi ini
    HANYA menerima PDF sesuai permintaan bisnis bahwa BA (Berita Acara)
    serah terima harus dalam bentuk PDF resmi.

    Return (filename, error). filename bernilai None jika error tidak kosong.
    File yang tidak diisi dianggap error juga (evidence wajib/mandatory).
    """
    if not file_storage or file_storage.filename == "":
        return None, "Evidence PDF wajib diupload."
    if not file_storage.filename.lower().endswith(".pdf"):
        return None, "Evidence wajib berupa file PDF (.pdf), format lain tidak diterima."
    filename = secure_filename(file_storage.filename)
    if not filename:
        return None, "Nama file tidak valid."
    unique_name = f"peminjaman_evidence_{datetime.now(WIB).strftime('%Y%m%d%H%M%S%f')}_{filename}"
    filepath = os.path.join(app.config["UPLOAD_FOLDER"], unique_name)
    try:
        file_storage.save(filepath)
        return unique_name, None
    except Exception as e:
        return None, f"Gagal menyimpan file: {str(e)}"


def gagal_dengan_form(endpoint, pesan, target, id_target=None, **url_kwargs):
    """Batalkan submit form (tambah/edit) tanpa menghilangkan data yang sudah
    diisi user. Data form disimpan sementara di session lalu dikirim balik
    ke halaman asal supaya modal yang sama otomatis terbuka lagi dengan
    field-field terisi seperti sebelumnya -- user tidak perlu mengetik ulang.

    `target` menandai modal mana yang harus dibuka ulang di halaman tujuan,
    mis. 'tambah_aset', 'edit_aset', atau 'tambah_peminjaman'.
    `id_target` (opsional) dipakai untuk modal edit yang butuh tahu ID data
    mana yang sedang diedit (mis. supaya openEditAset() tahu ID-nya).
    File upload (mis. foto/evidence) TIDAK bisa disimpan ulang oleh browser
    karena alasan keamanan, jadi user tetap perlu memilih ulang file-nya.
    """
    flash(pesan, "danger")
    session["form_repopulate"] = {
        "target": target,
        "id": id_target,
        "form": request.form.to_dict(),
    }
    return redirect(url_for(endpoint, **url_kwargs))


def convert_gdrive_to_thumbnail(url):
    if not url:
        return url
    if 'drive.google.com' not in url:
        return url
    match = re.search(r'/d/([a-zA-Z0-9_-]+)', url)
    if match:
        file_id = match.group(1)
        return f"https://drive.google.com/thumbnail?id={file_id}&sz=w1000"
    match = re.search(r'[?&]id=([a-zA-Z0-9_-]+)', url)
    if match:
        file_id = match.group(1)
        return f"https://drive.google.com/thumbnail?id={file_id}&sz=w1000"
    return url

def extract_link_from_cell(cell):
    """Ambil link Google Drive dari sebuah cell Excel.
    Prioritas: hyperlink yang nempel di cell (paling umum dipakai di file BA
    Transfer -- teks cell cuma nama file, link Drive-nya ada di hyperlink),
    baru kalau tidak ada, cek apakah teks cell itu sendiri sudah berupa URL."""
    if cell is None:
        return None
    if getattr(cell, "hyperlink", None) and cell.hyperlink.target:
        return cell.hyperlink.target
    value = cell.value
    if isinstance(value, str) and value.strip().lower().startswith(("http://", "https://")):
        return value.strip()
    return None


# Header yang dicari untuk sheet import Peminjaman (gaya "BA Transfer").
# key = versi ternormalisasi (lower, spasi/enter dirapikan) dari header Excel,
# value = nama field internal yang dipakai proses import.
PEMINJAMAN_HEADER_ALIASES = {
    "no": "no",
    "nama": "nama",
    "unit": "unit",
    "lokasi kerja": "lokasi_kerja",
    "jenis barang": "jenis_barang",
    "jenis transaksi": "jenis_transaksi",
    "tanggal awal": "tanggal_awal",
    "tanggal akhir": "tanggal_akhir",
    "keterangan": "keterangan",
    "evidence lampiran": "evidence_lampiran",
}


def _normalize_header(text):
    if text is None:
        return ""
    return " ".join(str(text).strip().lower().replace("\n", " ").split())


def find_peminjaman_sheet(wb):
    """Cari sheet + baris header yang paling cocok dengan format import
    Peminjaman (kolom: No, Nama, Unit, Lokasi Kerja, Jenis Barang,
    Jenis Transaksi, Tanggal Awal, Tanggal Akhir, Keterangan, Evidence
    Lampiran), berapa pun urutan kolomnya dan di sheet mana pun letaknya --
    tidak asumsi header selalu di baris 1 atau di sheet pertama.

    Return: (worksheet, header_row_index, {field: col_index}) atau
    (None, None, None) kalau tidak ketemu sheet yang cocok."""
    best = (None, None, None, 0)  # ws, header_row, col_map, jumlah_match

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        max_row_scan = min(ws.max_row or 1, 15)
        for row_idx in range(1, max_row_scan + 1):
            col_map = {}
            for col_idx in range(1, (ws.max_column or 1) + 1):
                header_text = _normalize_header(ws.cell(row=row_idx, column=col_idx).value)
                field = PEMINJAMAN_HEADER_ALIASES.get(header_text)
                if field and field not in col_map:
                    col_map[field] = col_idx
            # butuh minimal kolom-kolom inti supaya dianggap baris header yang valid
            wajib = {"nama", "jenis_transaksi", "evidence_lampiran"}
            if wajib.issubset(col_map.keys()) and len(col_map) > best[3]:
                best = (ws, row_idx, col_map, len(col_map))

    return best[0], best[1], best[2]


def parse_excel_date(value):
    """Parse nilai tanggal dari cell Excel yang bisa berupa datetime asli
    (kalau cell sudah diformat sebagai tanggal), angka serial Excel (kalau
    cell TIDAK diformat sebagai tanggal walau isinya tanggal -- ini terjadi
    di beberapa baris file BA Transfer), atau string tanggal biasa."""
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            base = datetime(1899, 12, 30)
            return (base + timedelta(days=float(value))).date()
        except Exception:
            return None
    if isinstance(value, str):
        v = value.strip()
        if not v:
            return None
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(v, fmt).date()
            except ValueError:
                continue
        return None
    return None


def catat_aktivitas(aksi, target_model, target_id, deskripsi=None, data_lama=None, data_baru=None):
    """Catat aktivitas admin ke tabel aktivitas_log."""
    log = AktivitasLog(
        id_user=current_user.id,
        aksi=aksi,
        target_model=target_model,
        target_id=target_id,
        deskripsi=deskripsi,
        data_lama=data_lama,
        data_baru=data_baru
    )
    db.session.add(log)


# Urutan & label field aset yang ditampilkan di detail histori (Tambah/Edit/Hapus).
# Tambahkan/rename di sini kalau ada field baru -> otomatis ikut muncul di histori.
FIELD_LABELS_ASET = [
    ("kode_aset", "Kode Aset"),
    ("nama", "Nama Aset"),
    ("foto_url", "Foto"),
    ("kategori", "Kategori"),
    ("tipe_aset", "Tipe Aset"),
    ("area", "Area"),
    ("fungsi", "Fungsi"),
    ("merek", "Merek"),
    ("serial_number", "Serial Number"),
    ("spesifikasi", "Spesifikasi"),
    ("volume", "Volume"),
    ("satuan", "Satuan"),
    ("status", "Status"),
    ("gedung", "Gedung"),
    ("lantai", "Lantai"),
    ("ruangan", "Ruangan"),
    ("link_qr", "Link QR"),
    ("tanggal_datang", "Tanggal Datang"),
    ("keterangan", "Keterangan"),
]


def snapshot_aset(aset, kategori_nama=None):
    """Snapshot lengkap satu aset untuk disimpan di log histori (CREATE/UPDATE/DELETE).
    kategori_nama dioper manual (bukan lewat relationship) supaya tidak kena masalah
    cache relasi SQLAlchemy saat kategori baru saja diganti dalam request yang sama."""
    return {
        "kode_aset": aset.kode_aset,
        "nama": aset.nama,
        "foto_url": aset.foto_url,
        "kategori": kategori_nama,
        "tipe_aset": aset.tipe_aset,
        "area": aset.area,
        "fungsi": aset.fungsi,
        "merek": aset.merek,
        "serial_number": aset.serial_number,
        "spesifikasi": aset.spesifikasi,
        "volume": aset.volume,
        "satuan": aset.satuan,
        "status": aset.status_aset,
        "gedung": aset.gedung,
        "lantai": aset.lantai,
        "ruangan": aset.ruangan,
        "link_qr": aset.link_qr,
        "tanggal_datang": aset.tanggal_datang.strftime("%Y-%m-%d") if aset.tanggal_datang else None,
        "keterangan": aset.keterangan,
    }


# Urutan & label field peminjaman yang ditampilkan di detail histori
# (perpanjangan, pengembalian, upload evidence, dsb). Field "evidence" &
# "foto_url" otomatis dikenali sebagai file (gambar/PDF) -- lihat FILE_FIELD_KEYS.
FIELD_LABELS_PEMINJAMAN = [
    ("nama_peminjam", "Nama Peminjam"),
    ("unit", "Unit"),
    ("lokasi_kerja", "Lokasi Kerja"),
    ("jenis_transaksi", "Jenis Transaksi"),
    ("barang", "Barang / Aset Dipinjam"),
    ("tanggal_pinjam", "Tanggal Pinjam"),
    ("tanggal_rencana_kembali", "Rencana Tanggal Kembali"),
    ("tanggal_dikembalikan", "Tanggal Dikembalikan"),
    ("status", "Status"),
    ("status_perpanjangan", "Status Perpanjangan"),
    ("evidence", "Evidence / Bukti (BA Serah Terima)"),
    ("keterangan_evidence", "Keterangan Evidence"),
]

# Field yang isinya nama file (gambar atau dokumen/PDF) hasil upload --
# ditampilkan sebagai preview foto (kalau gambar) atau link "Lihat PDF"
# (kalau dokumen), bukan sekadar teks nama file.
FILE_FIELD_KEYS = {"foto_url", "evidence", "foto"}


def snapshot_peminjaman(p):
    """Snapshot lengkap satu peminjaman untuk disimpan di log histori
    (dipakai saat DELETE supaya datanya tidak hilang begitu saja dari History)."""
    barang = [pa.aset.nama for pa in p.aset_terkait if pa.aset]
    return {
        "nama_peminjam": p.nama_peminjam,
        "unit": p.unit,
        "lokasi_kerja": p.lokasi_kerja,
        "jenis_transaksi": p.jenis_transaksi,
        "barang": barang,
        "tanggal_pinjam": p.tanggal_pinjam.strftime("%Y-%m-%d") if p.tanggal_pinjam else None,
        "tanggal_rencana_kembali": p.tanggal_rencana_kembali.strftime("%Y-%m-%d") if p.tanggal_rencana_kembali else None,
        "tanggal_dikembalikan": p.tanggal_dikembalikan.strftime("%Y-%m-%d") if p.tanggal_dikembalikan else None,
        "status": p.status,
        "status_perpanjangan": p.status_perpanjangan,
        "evidence": p.evidence,
    }


# Urutan & label field maintenance yang ditampilkan di detail histori
# (Tambah/Edit/Hapus jadwal maintenance).
FIELD_LABELS_MAINTENANCE = [
    ("aset", "Aset"),
    ("kode_aset", "Kode Aset"),
    ("kategori", "Kategori"),
    ("judul", "Judul Maintenance"),
    ("deskripsi", "Deskripsi"),
    ("vendor", "Vendor"),
    ("tipe", "Tipe Maintenance"),
    ("tanggal_mulai", "Tanggal Mulai"),
    ("tanggal_akhir", "Tanggal Akhir"),
    ("biaya", "Biaya"),
    ("status", "Status"),
]


def snapshot_maintenance(m):
    """Snapshot lengkap satu jadwal maintenance untuk disimpan di log histori
    (CREATE/UPDATE/DELETE), supaya "Detail Aktivitas" bisa menampilkan semua
    field seperti halaman Detail Aset, bukan cuma sebagian."""
    aset = m.aset
    return {
        "aset": aset.nama if aset else None,
        "kode_aset": aset.kode_aset if aset else None,
        "kategori": m.kategori,
        "judul": m.judul,
        "deskripsi": m.deskripsi,
        "vendor": m.vendor,
        "tipe": m.tipe,
        "tanggal_mulai": m.tanggal_mulai.strftime("%Y-%m-%d") if m.tanggal_mulai else None,
        "tanggal_akhir": m.tanggal_akhir.strftime("%Y-%m-%d") if m.tanggal_akhir else None,
        "biaya": float(m.biaya) if m.biaya is not None else None,
        "status": m.status,
    }


def label_aktivitas(a):
    """Tentukan label aktivitas yang ditampilkan di History / Detail Aktivitas,
    disesuaikan per target_model DAN per jenis perubahan supaya lebih informatif
    (mis. "Perpanjangan Peminjaman" bukan cuma "Update Peminjaman", "Upload Foto
    Dokumentasi" bukan cuma "Edit Maintenance"). Dipakai bersama oleh history_list
    dan aktivitas_detail supaya labelnya selalu konsisten di kedua halaman."""
    data_baru = a.data_baru or {}
    data_lama = a.data_lama or {}

    if a.target_model == "Aset":
        return {
            "CREATE": "Tambah Aset",
            "UPDATE": "Edit Aset",
            "DELETE": "Hapus Aset",
            "MOVE": "Pemindahan Aset",
        }.get(a.aksi, a.aksi)

    if a.target_model == "Peminjaman":
        if a.aksi == "CREATE":
            return "Peminjaman Aset"
        if a.aksi == "DELETE":
            return "Hapus Peminjaman"
        if a.aksi == "IMPORT":
            return "Import Data Peminjaman"
        if a.aksi == "UPDATE":
            keys = set(data_baru.keys())
            if data_baru.get("status") == "Dikembalikan":
                return "Pengembalian Aset"
            if data_baru.get("status_perpanjangan") == "Tidak Diperpanjang":
                return "Konfirmasi Tidak Diperpanjang"
            if "tanggal_rencana_kembali" in keys:
                return "Perpanjangan Peminjaman"
            if "evidence" in keys:
                return "Upload Evidence Peminjaman"
            return "Update Peminjaman"
        return a.aksi

    if a.target_model == "Maintenance":
        is_foto = ("foto" in data_baru) or ("foto" in data_lama)
        if a.aksi == "CREATE":
            return "Tambah Maintenance"
        if a.aksi == "DELETE":
            return "Hapus Maintenance"
        if a.aksi == "UPDATE" and is_foto:
            return "Hapus Foto Dokumentasi" if not data_baru.get("foto") else "Upload Foto Dokumentasi"
        if a.aksi == "UPDATE":
            return "Edit Jadwal Maintenance"
        return a.aksi

    return {"CREATE": "Tambah Data", "UPDATE": "Update Data", "DELETE": "Hapus Data"}.get(a.aksi, a.aksi)


def build_field_diff(data_lama, data_baru, field_labels=None):
    """Bandingkan dua snapshot (aset/peminjaman/dll) field-per-field, hasilkan
    daftar siap-tampil (dipakai di halaman Detail Aktivitas / History) mengikuti
    urutan `field_labels` (default FIELD_LABELS_ASET), plus field lain yang
    mungkin ada di data tapi belum terdaftar di label map."""
    field_labels = field_labels or FIELD_LABELS_ASET
    data_lama = data_lama or {}
    data_baru = data_baru or {}
    known_keys = [k for k, _ in field_labels]
    extra_keys = [k for k in {**data_lama, **data_baru}.keys() if k not in known_keys]
    all_fields = field_labels + [(k, k.replace("_", " ").title()) for k in extra_keys]

    diff = []
    for key, label in all_fields:
        if key not in data_lama and key not in data_baru:
            continue
        old_val = data_lama.get(key)
        new_val = data_baru.get(key)
        diff.append({
            "key": key,
            "label": label,
            "old": old_val,
            "new": new_val,
            "changed": old_val != new_val,
            "is_file": key in FILE_FIELD_KEYS,
        })
    return diff

def role_required(*roles):
    def decorator(f):
        @wraps(f)
        def wrapped(*args, **kwargs):
            if not current_user.is_authenticated:
                return redirect(url_for("login"))
            if current_user.role not in roles:
                abort(403)
            return f(*args, **kwargs)
        return wrapped
    return decorator

@app.route("/api/aset/search")
@login_required
def api_aset_search():
    """API untuk mencari aset berdasarkan nama atau kode aset (auto-complete)."""
    q = request.args.get("q", "").strip()
    if len(q) < 2:
        return jsonify([])
    
    hasil = Aset.query.filter(
        db.or_(
            Aset.nama.ilike(f"%{q}%"),
            Aset.kode_aset.ilike(f"%{q}%")
        )
    ).limit(15).all()
    
    data = []
    for a in hasil:
        data.append({
            "id": a.id,
            "kode": a.kode_aset,
            "nama": a.nama,
            "kategori": a.kategori_ref.nama if a.kategori_ref else "Tidak ada kategori"
        })
    return jsonify(data)

def catat_log(tiket, status_lama, status_baru):
    db.session.add(LogStatus(
        id_tiket=tiket.id,
        status_lama=status_lama,
        status_baru=status_baru,
        id_user_pengubah=current_user.id
    ))


@app.errorhandler(403)
def forbidden(e):
    return render_template("403.html"), 403


@app.errorhandler(IntegrityError)
def handle_integrity_error(e):
    db.session.rollback()
    flash(
        "Data tidak bisa disimpan: kemungkinan data duplikat atau masih "
        "terkait dengan data lain.",
        "danger",
    )
    return redirect(request.referrer or url_for("dashboard"))


# ---------------------------------------------------------------------------
# AUTH
# ---------------------------------------------------------------------------
@app.route("/login", methods=["GET", "POST"])
@limiter.limit("8 per minute", methods=["POST"])
def login():
    if current_user.is_authenticated:
        return redirect(url_for("dashboard"))
    if request.method == "POST":
        email = request.form.get("email", "").strip()
        password = request.form.get("password", "")
        user = User.query.filter_by(email=email).first()
        if user and check_password_hash(user.password, password):
            if not user.is_active:
                flash("Akun ini sudah dinonaktifkan. Hubungi admin.", "danger")
                return render_template("login.html")
            if user.banned_until and user.banned_until > datetime.now(WIB).replace(tzinfo=None):
                pesan_ban = (
                    f"Akun ini sedang di-ban sementara sampai "
                    f"{user.banned_until.strftime('%d-%m-%Y %H:%M')} WIB."
                )
                if user.ban_reason:
                    pesan_ban += f" Alasan: {user.ban_reason}."
                pesan_ban += " Hubungi admin."
                flash(pesan_ban, "danger")
                return render_template("login.html")
            login_user(user)
            return redirect(url_for("dashboard"))
        flash("Email atau password salah.", "danger")
    return render_template("login.html")


@app.route("/logout")
@login_required
def logout():
    logout_user()
    return redirect(url_for("login"))


# ---------------------------------------------------------------------------
# DASHBOARD
# ---------------------------------------------------------------------------
@app.route("/")
@login_required
def dashboard():
    total_aset = Aset.query.count()
    total_rusak = Aset.query.filter_by(status_aset="Rusak").count()
    total_kategori = Kategori.query.count()
    
    # Untuk user: statistik CAPEX/OPEX
    total_capex = Aset.query.filter_by(tipe_aset="CAPEX").count()
    total_opex = Aset.query.filter_by(tipe_aset="OPEX").count()

    total_maintenance = Maintenance.query.count()
    total_pemindahan = Tiket.query.filter_by(jenis_tiket="Pemindahan").count()
    
    # Kategori terbanyak
    kategori_terbanyak = None
    if total_aset > 0:
        kategori_terbanyak = db.session.query(
            Kategori.nama, db.func.count(Aset.id).label('jumlah')
        ).outerjoin(Aset, Aset.id_kategori == Kategori.id)\
         .group_by(Kategori.id)\
         .order_by(db.desc('jumlah'))\
         .first()
        if kategori_terbanyak:
            kategori_terbanyak = {
                "nama": kategori_terbanyak[0],
                "jumlah": kategori_terbanyak[1]
            }
    
    # Grafik kategori
    kategori_chart = (
        db.session.query(Kategori.nama, db.func.count(Aset.id))
        .outerjoin(Aset, Aset.id_kategori == Kategori.id)
        .group_by(Kategori.id)
        .all()
    )
    chart_labels = [k[0] for k in kategori_chart]
    chart_values = [k[1] for k in kategori_chart]
    
    # History terbaru (hanya untuk admin)
    history_terbaru = []
    if current_user.role == ROLE_ADMIN:
        # Ambil semua event (tiket + aktivitas log) untuk admin
        events = []
        for t in Tiket.query.order_by(Tiket.created_at.desc()).limit(10).all():
            events.append({
                "id": t.id,
                "jenis_tiket": t.jenis_tiket,
                "nama_pemohon": t.nama_pemohon,
                "catatan": t.catatan[:50] if t.catatan else "-",
                "created_at": t.created_at,
                "is_tiket": True
            })
        
        # Gabungkan dengan aktivitas log (urutkan berdasarkan waktu)
        from models import AktivitasLog
        for a in AktivitasLog.query.order_by(AktivitasLog.created_at.desc()).limit(5).all():
            user = User.query.get(a.id_user)
            events.append({
                "id": a.id,
                "jenis_tiket": "Aktivitas",
                "nama_pemohon": user.name if user else "System",
                "catatan": a.deskripsi[:50] if a.deskripsi else "-",
                "created_at": a.created_at,
                "is_tiket": False
            })
        
        # Urutkan berdasarkan waktu terbaru dan ambil 5 teratas
        events.sort(key=lambda x: x["created_at"], reverse=True)
        history_terbaru = events[:5]

    # History Maintenance terbaru (5 jadwal terbaru)
    maintenance_terbaru = (
        Maintenance.query.order_by(Maintenance.tanggal_mulai.desc(), Maintenance.id.desc())
        .limit(5)
        .all()
    )

    # History Pemindahan terbaru (5 tiket pemindahan terbaru)
    pemindahan_terbaru = (
        Tiket.query.filter_by(jenis_tiket="Pemindahan")
        .order_by(Tiket.created_at.desc())
        .limit(5)
        .all()
    )

    # ================================================================
    # KALENDER PEMINJAMAN + NOTIFIKASI PERPANJANGAN H-10
    # ================================================================
    today = datetime.now(WIB).date()
    batas_notif = today + timedelta(days=10)

    # Peminjaman yang butuh konfirmasi perpanjangan: rencana kembali
    # sudah lewat (terlambat) ATAU akan jatuh tempo dalam 10 hari ke depan,
    # dan belum pernah dikonfirmasi (status_perpanjangan masih kosong).
    peminjaman_reminder = []
    peminjaman_calendar = {}
    if current_user.role == ROLE_ADMIN:
        peminjaman_reminder = (
            Peminjaman.query.filter(
                Peminjaman.status == "Dipinjam",
                Peminjaman.tanggal_rencana_kembali.isnot(None),
                Peminjaman.tanggal_rencana_kembali <= batas_notif,
                Peminjaman.status_perpanjangan.is_(None),
            )
            .order_by(Peminjaman.tanggal_rencana_kembali.asc())
            .all()
        )

        # Data untuk ditandai di kalender: semua peminjaman aktif yang punya tanggal rencana kembali
        peminjaman_aktif = Peminjaman.query.filter(
            Peminjaman.status == "Dipinjam",
            Peminjaman.tanggal_rencana_kembali.isnot(None),
        ).all()
        for p in peminjaman_aktif:
            key = p.tanggal_rencana_kembali.strftime("%Y-%m-%d")
            aset_names = ", ".join([pa.aset.nama for pa in p.aset_terkait if pa.aset]) or "-"
            peminjaman_calendar.setdefault(key, []).append({
                "id": p.id,
                "nama_peminjam": p.nama_peminjam,
                "unit": p.unit or "-",
                "lokasi_kerja": p.lokasi_kerja or "-",
                "aset": aset_names,
                "tanggal_pinjam": p.tanggal_pinjam.strftime("%d-%m-%Y") if p.tanggal_pinjam else "-",
                "tanggal_rencana_kembali": p.tanggal_rencana_kembali.strftime("%d-%m-%Y") if p.tanggal_rencana_kembali else "-",
                "keterangan": p.keterangan or "-",
                "status": p.status,
                "overdue": p.tanggal_rencana_kembali < today,
                "butuh_konfirmasi": p.status_perpanjangan is None,
            })

    # Daftar hari libur nasional 2026 (SKB 3 Menteri) untuk ditandai di kalender
    hari_libur_nasional = {
        "2026-01-01": "Tahun Baru Masehi",
        "2026-01-16": "Isra Mikraj",
        "2026-02-17": "Tahun Baru Imlek",
        "2026-03-19": "Hari Suci Nyepi",
        "2026-03-21": "Idulfitri (Hari 1)",
        "2026-03-22": "Idulfitri (Hari 2)",
        "2026-04-03": "Wafat Yesus Kristus",
        "2026-04-05": "Paskah",
        "2026-05-01": "Hari Buruh",
        "2026-05-14": "Kenaikan Yesus Kristus",
        "2026-05-27": "Iduladha",
        "2026-05-31": "Hari Raya Waisak",
        "2026-06-01": "Hari Lahir Pancasila",
        "2026-06-16": "1 Muharam",
        "2026-08-17": "Proklamasi Kemerdekaan",
        "2026-08-25": "Maulid Nabi Muhammad",
        "2026-12-25": "Hari Raya Natal",
    }

    return render_template(
        "dashboard.html",
        total_aset=total_aset,
        total_rusak=total_rusak,
        total_kategori=total_kategori,
        total_maintenance=total_maintenance,  
        total_pemindahan=total_pemindahan,
        total_capex=total_capex,
        total_opex=total_opex,
        kategori_terbanyak=kategori_terbanyak,
        chart_labels=chart_labels,
        chart_values=chart_values,
        history_terbaru=history_terbaru,
        maintenance_terbaru=maintenance_terbaru,
        pemindahan_terbaru=pemindahan_terbaru,
        peminjaman_reminder=peminjaman_reminder,
        peminjaman_calendar_json=json.dumps(peminjaman_calendar),
        hari_libur_json=json.dumps(hari_libur_nasional),
        calendar_year=today.year,
        today=today,
        today_str=today.strftime("%Y-%m-%d"),
    )

@app.route("/register", methods=["GET", "POST"])
def register():
    if current_user.is_authenticated:
        return redirect(url_for("dashboard"))
    
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        email = request.form.get("email", "").strip()
        password = request.form.get("password", "")
        confirm_password = request.form.get("confirm_password", "")
        
        # Validasi
        if not name or not email or not password:
            flash("Semua field wajib diisi.", "danger")
            return render_template("register.html")
        
        # Validasi email harus @gmail.com
        if not email.endswith("@gmail.com"):
            flash("Hanya email dengan domain @gmail.com yang diperbolehkan.", "danger")
            return render_template("register.html")
        
        # Validasi email sudah terdaftar
        if User.query.filter_by(email=email).first():
            flash("Email sudah terdaftar. Silakan login.", "danger")
            return render_template("register.html")
        
        # Validasi password minimal 8 karakter
        if len(password) < 8:
            flash("Password minimal 8 karakter.", "danger")
            return render_template("register.html")
        
        # Validasi konfirmasi password
        if password != confirm_password:
            flash("Password dan konfirmasi password tidak cocok.", "danger")
            return render_template("register.html")
        
        # Buat user baru dengan role 'user'
        user = User(
            name=name,
            email=email,
            password=generate_password_hash(password),
            role="user",  # default role user
            is_active=True
        )
        db.session.add(user)
        db.session.commit()
        
        flash("Akun berhasil dibuat! Silakan login.", "success")
        return redirect(url_for("login"))
    
    return render_template("register.html")

# ---------------------------------------------------------------------------
# ASET (CRUD)
# ---------------------------------------------------------------------------
JENIS_ASET_OPTIONS = ["Operasional", "Pusat"]

# Nama kota/lokasi untuk tiap kode area TCU, ditampilkan di belakang kode area
# pada label dropdown Gedung (mis. "TCU1" -> "TCU1 Jakarta").
AREA_LOKASI_MAP = {
    "TCU1": "Jakarta",
    "TCU2": "Bandung",
    "TCU3": "Semarang",
    "TCU4": "Makassar",
}


def format_area_label(area):
    """Tambahkan nama kota di belakang kode area, mis. 'TCU1' -> 'TCU1 Jakarta'.
    Kode area yang tidak ada di AREA_LOKASI_MAP ditampilkan apa adanya."""
    if not area:
        return area
    kota = AREA_LOKASI_MAP.get(area)
    return f"{area} {kota}" if kota else area


# Daftarkan sebagai fungsi global Jinja supaya bisa dipanggil langsung dari
# template, mis. {{ format_area_label(a.area) }} -- dipakai di kolom Lokasi
# tabel Data Aset supaya nama kota TCU juga tampil di sana, bukan cuma di
# dropdown filter.
app.jinja_env.globals["format_area_label"] = format_area_label


def is_file_gambar(filename):
    """True kalau nama file yang disimpan (mis. hasil upload evidence/foto)
    berekstensi gambar. Dipakai di template history/detail supaya file
    gambar ditampilkan sebagai preview <img>, sedangkan PDF/Word ditampilkan
    sebagai link/tombol unduh biasa."""
    if not filename or "." not in filename:
        return False
    return filename.rsplit(".", 1)[1].lower() in ALLOWED_EXT


# Dipakai dari template, mis. {% if is_file_gambar(d.new) %}
app.jinja_env.globals["is_file_gambar"] = is_file_gambar


def parse_gedung_value(value):
    """
    Parse nilai filter 'gedung' yang dikirim dari form.

    Sejak perbaikan bug "Filter Lanjutan Gedung", value dropdown Gedung di
    halaman Data Aset dikirim dalam format gabungan "AREA||NAMA_GEDUNG"
    (bukan cuma nama gedung), karena banyak gedung punya NAMA YANG SAMA
    di beberapa area/TCU berbeda (mis. "Gedung D" ada di TCU1, TCU2, dan
    TCU3 sekaligus). Kalau cuma difilter pakai nama gedung saja, data dari
    ke-3 TCU itu ikut tercampur, dan opsi <select> yang value-nya sama-sama
    "Gedung D" bikin browser salah menampilkan pilihan yang ter-select
    (efeknya: pilih TCU2 tapi yang muncul kepilih TCU3).

    Fungsi ini tetap kompatibel dengan pemanggil lama (form Tiket, Riwayat,
    Maintenance) yang masih mengirim nama gedung polos tanpa "||" -- dalam
    kasus itu, area dikembalikan kosong dan filter area diabaikan (perilaku
    lama tetap jalan, tidak ada yang rusak).
    """
    if not value:
        return "", ""
    if "||" in value:
        area_part, gedung_part = value.split("||", 1)
        return area_part, gedung_part
    return "", value


def upsert_lokasi_master(area, gedung, lantai, ruangan):
    """Simpan/perbarui data Area, Gedung, Lantai, dan Ruangan ke tabel master
    (lihat models.py: Area, Gedung, Lantai, Ruangan) berdasarkan satu baris
    data aset (biasanya dipanggil per-baris saat import Excel di
    /aset/import). Tabel-tabel ini TIDAK ditampilkan langsung ke user --
    hanya dipakai sebagai sumber data dropdown Area/Gedung/Lantai/Ruangan di
    form Tambah & Edit Aset.

    Kalau kombinasi data sudah ada, tidak dibuat duplikat (get-or-create).
    Nama gedung boleh sama di area yang berbeda (mis. "Gedung D" di TCU1 dan
    TCU2), jadi Gedung dicocokkan berdasarkan kombinasi nama + area.
    """
    area = (area or "").strip()
    gedung = (gedung or "").strip()
    lantai = (lantai or "").strip()
    ruangan = (ruangan or "").strip()

    area_obj = None
    if area:
        area_obj = Area.query.filter(db.func.lower(Area.nama) == area.lower()).first()
        if not area_obj:
            area_obj = Area(nama=area)
            db.session.add(area_obj)
            db.session.flush()

    gedung_obj = None
    # "-" dipakai sebagai placeholder gedung kosong di proses import (lihat
    # aset_import) -- jangan ikut dimasukkan sebagai data master.
    if gedung and gedung != "-":
        q = Gedung.query.filter(db.func.lower(Gedung.nama) == gedung.lower())
        q = q.filter_by(id_area=area_obj.id) if area_obj else q.filter(Gedung.id_area.is_(None))
        gedung_obj = q.first()
        if not gedung_obj:
            gedung_obj = Gedung(nama=gedung, id_area=area_obj.id if area_obj else None)
            db.session.add(gedung_obj)
            db.session.flush()

    lantai_obj = None
    if lantai and gedung_obj:
        lantai_obj = Lantai.query.filter(
            db.func.lower(Lantai.nama) == lantai.lower(), Lantai.id_gedung == gedung_obj.id
        ).first()
        if not lantai_obj:
            lantai_obj = Lantai(nama=lantai, id_gedung=gedung_obj.id)
            db.session.add(lantai_obj)
            db.session.flush()

    if ruangan and ruangan != "-" and gedung_obj:
        q = Ruangan.query.filter(
            db.func.lower(Ruangan.nama) == ruangan.lower(), Ruangan.id_gedung == gedung_obj.id
        )
        q = q.filter_by(id_lantai=lantai_obj.id) if lantai_obj else q.filter(Ruangan.id_lantai.is_(None))
        ruangan_obj = q.first()
        if not ruangan_obj:
            ruangan_obj = Ruangan(
                nama=ruangan, id_gedung=gedung_obj.id, id_lantai=lantai_obj.id if lantai_obj else None
            )
            db.session.add(ruangan_obj)


def build_lokasi_master():
    """Bangun data master lokasi (Area -> Gedung -> Lantai -> Ruangan) untuk
    dropdown berjenjang. Dipakai di form Tambah/Edit Aset (aset_list) dan
    juga di form Pemindahan Aset per-item (aset_pemindahan_form)."""
    return {
        "area": [
            {"id": a.id, "nama": a.nama, "label": format_area_label(a.nama)}
            for a in Area.query.order_by(Area.nama).all()
        ],
        "gedung": [
            {"id": g.id, "nama": g.nama, "id_area": g.id_area}
            for g in Gedung.query.order_by(Gedung.nama).all()
        ],
        "lantai": [
            {"id": l.id, "nama": l.nama, "id_gedung": l.id_gedung}
            for l in Lantai.query.order_by(Lantai.nama).all()
        ],
        "ruangan": [
            {"id": r.id, "nama": r.nama, "id_gedung": r.id_gedung, "id_lantai": r.id_lantai}
            for r in Ruangan.query.order_by(Ruangan.nama).all()
        ],
    }


@app.route("/aset")
@login_required
def aset_list():
    q = request.args.get("q", "").strip()
    status = request.args.get("status", "")
    kategori_id = request.args.get("kategori", "")
    gedung = request.args.get("gedung", "")  # format: "AREA||NAMA_GEDUNG"
    lantai = request.args.get("lantai", "")
    ruangan = request.args.get("ruangan", "")
    tipe = request.args.get("tipe", "")

    gedung_area, gedung_nama = parse_gedung_value(gedung)

    query = Aset.query
    if q:
        query = query.filter(
            db.or_(Aset.nama.ilike(f"%{q}%"), Aset.kode_aset.ilike(f"%{q}%"))
        )
    if status:
        query = query.filter_by(status_aset=status)
    if kategori_id:
        query = query.filter_by(id_kategori=kategori_id)
    if gedung:
        # PERBAIKAN: filter berdasarkan area + nama gedung, bukan nama gedung
        # saja -- karena beberapa nama gedung (mis. "Gedung D") dipakai di
        # lebih dari satu area/TCU. Filter by nama gedung saja akan
        # menggabungkan data dari area yang berbeda-beda.
        query = query.filter_by(gedung=gedung_nama)
        if gedung_area:
            query = query.filter_by(area=gedung_area)
    if lantai:
        query = query.filter_by(lantai=lantai)
    if ruangan:
        query = query.filter_by(ruangan=ruangan)
    if tipe:
        query = query.filter_by(tipe_aset=tipe)

    filter_aktif = bool(q or status or kategori_id or gedung or lantai or ruangan or tipe)

    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 10, type=int)
    if per_page not in [10, 25, 50, 100]:
        per_page = 10

    pagination = query.order_by(Aset.id.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )
    daftar_aset = pagination.items
    kategori_all = Kategori.query.all()
    
    # --- TAMBAHAN: Ambil daftar AREA + GEDUNG (unik) ---
    gedung_all = (
        db.session.query(Aset.area, Aset.gedung)
        .filter(Aset.gedung.isnot(None), Aset.gedung != "")
        .distinct()
        .order_by(Aset.area, Aset.gedung)
        .all()
    )
    # Format: "Area - Gedung". Value dibuat UNIK per kombinasi area+gedung
    # ("AREA||NAMA_GEDUNG"), bukan nama gedung saja -- lihat parse_gedung_value()
    # untuk penjelasan kenapa ini penting (nama gedung bisa dobel di area lain).
    gedung_all_formatted = [
        {
            "value": f"{g.area}||{g.gedung}" if g.area else g.gedung,
            "label": f"{format_area_label(g.area)} - {g.gedung}" if g.area else g.gedung,
        }
        for g in gedung_all
    ]

    total_keseluruhan = Aset.query.count()

    # Kalau sebelumnya ada submit form Tambah/Edit Aset yang gagal validasi,
    # ambil (dan hapus) data yang sempat disimpan supaya modal yang sama bisa
    # otomatis terbuka lagi dengan field-field yang sudah terisi -- user
    # tidak perlu mengetik ulang semuanya. Lihat helper gagal_dengan_form().
    repop = session.pop("form_repopulate", None)

    # --- TAMBAHAN: data master lokasi untuk dropdown berjenjang Area ->
    # Gedung -> Lantai -> Ruangan di form Tambah/Edit Aset. Diambil dari
    # tabel Area/Gedung/Lantai/Ruangan (terisi otomatis lewat import Excel,
    # lihat upsert_lokasi_master()). Dikirim sebagai JSON supaya JS bisa
    # melakukan filter berjenjang tanpa perlu request tambahan ke server.
    lokasi_master = build_lokasi_master()

    # --- TAMBAHAN: opsi dropdown Fungsi Barang & Satuan. Field ini TIDAK
    # punya tabel master baru (sesuai permintaan) -- opsinya cukup diambil
    # dari nilai unik yang sudah ada di tabel Aset.
    fungsi_all = [
        r[0] for r in db.session.query(Aset.fungsi)
        .filter(Aset.fungsi.isnot(None), Aset.fungsi != "")
        .distinct().order_by(Aset.fungsi).all()
    ]
    satuan_all = [
        r[0] for r in db.session.query(Aset.satuan)
        .filter(Aset.satuan.isnot(None), Aset.satuan != "")
        .distinct().order_by(Aset.satuan).all()
    ]

    return render_template(
        "aset/list.html",
        daftar_aset=daftar_aset,
        kategori_all=kategori_all,
        pagination=pagination,
        gedung_all=gedung_all_formatted,  # <-- KIRIM FORMAT BARU
        filter_aktif=filter_aktif,
        total_keseluruhan=total_keseluruhan,
        repop=repop,
        lokasi_master=lokasi_master,
        fungsi_all=fungsi_all,
        satuan_all=satuan_all,
    )


@app.route("/api/lantai")
@login_required
def api_lantai():
    gedung_raw = request.args.get("gedung", "")
    if not gedung_raw:
        return jsonify([])
    gedung_area, gedung_nama = parse_gedung_value(gedung_raw)
    filters = [Aset.gedung == gedung_nama, Aset.lantai.isnot(None), Aset.lantai != ""]
    if gedung_area:
        filters.append(Aset.area == gedung_area)
    hasil = (
        db.session.query(Aset.lantai)
        .filter(*filters)
        .distinct()
        .order_by(Aset.lantai)
        .all()
    )
    return jsonify([r[0] for r in hasil])


@app.route("/api/ruangan")
@login_required
def api_ruangan():
    gedung_raw = request.args.get("gedung", "")
    lantai = request.args.get("lantai", "")
    if not gedung_raw:
        return jsonify([])
    gedung_area, gedung_nama = parse_gedung_value(gedung_raw)
    filters = [Aset.gedung == gedung_nama, Aset.ruangan.isnot(None), Aset.ruangan != ""]
    if gedung_area:
        filters.append(Aset.area == gedung_area)
    if lantai:
        filters.append(Aset.lantai == lantai)
    hasil = (
        db.session.query(Aset.ruangan)
        .filter(*filters)
        .distinct()
        .order_by(Aset.ruangan)
        .all()
    )
    return jsonify([r[0] for r in hasil])


@app.route("/api/aset-by-lokasi")
@login_required
def api_aset_by_lokasi():
    gedung_raw = request.args.get("gedung", "")
    lantai = request.args.get("lantai", "")
    ruangan = request.args.get("ruangan", "")
    if not gedung_raw:
        return jsonify([])
    gedung_area, gedung_nama = parse_gedung_value(gedung_raw)
    filters = [Aset.gedung == gedung_nama]
    if gedung_area:
        filters.append(Aset.area == gedung_area)
    if lantai:
        filters.append(Aset.lantai == lantai)
    if ruangan:
        filters.append(Aset.ruangan == ruangan)
    hasil = Aset.query.filter(*filters).all()
    return jsonify([{
        "id": a.id,
        "kode": a.kode_aset,
        "nama": a.nama,
        "kategori": a.kategori_ref.nama if a.kategori_ref else ""  # <-- NAMA KATEGORI ASET
    } for a in hasil])


@app.route("/api/aset-by-kategori/<int:kategori_id>")
@login_required
def api_aset_by_kategori(kategori_id):
    """Daftar aset milik satu kategori (dipakai chained dropdown Kategori -> Aset)."""
    hasil = Aset.query.filter_by(id_kategori=kategori_id).order_by(Aset.nama).all()
    return jsonify([
        {
            "id": a.id,
            "kode": a.kode_aset,
            "nama": a.nama,
            "kategori": a.kategori_ref.nama if a.kategori_ref else "",
        }
        for a in hasil
    ])

@app.route("/aset/create", methods=["POST"])
@login_required
@role_required(ROLE_ADMIN)
def aset_create():
    kode_aset = request.form.get("kode_aset", "").strip()
    if Aset.query.filter_by(kode_aset=kode_aset).first():
        return gagal_dengan_form(
            "aset_list", "Kode aset sudah digunakan.", "tambah_aset"
        )

    # Ambil semua field
    area = request.form.get("area", "").strip() or None
    nama = request.form.get("nama", "").strip()
    fungsi = request.form.get("fungsi", "").strip() or None
    jenis_barang = request.form.get("jenis_barang", "").strip()  # nama kategori
    merek = request.form.get("merek", "").strip() or None
    serial_number = request.form.get("serial_number", "").strip() or None
    spesifikasi = request.form.get("spesifikasi", "").strip() or None
    tipe_aset = request.form.get("tipe_aset", "").strip() or None
    volume = request.form.get("volume", "").strip() or None
    satuan = request.form.get("satuan", "").strip() or None
    status_aset = request.form.get("status_aset", "Baik")
    gedung = request.form.get("gedung", "").strip()
    ruangan = request.form.get("ruangan", "").strip()
    lantai = request.form.get("lantai", "").strip() or None
    link_qr = request.form.get("link_qr", "").strip() or None
    tanggal_datang_str = request.form.get("tanggal_datang", "").strip()
    tanggal_datang = date.today()  # default ke tanggal hari ini kalau kosong
    if tanggal_datang_str:
        try:
            tanggal_datang = datetime.strptime(tanggal_datang_str, "%Y-%m-%d").date()
        except ValueError:
            pass
    keterangan = request.form.get("keterangan", "").strip() or None

    # Jenis Barang (Kategori) sekarang WAJIB diisi -- aset tidak boleh
    # tersimpan tanpa kategori sama sekali.
    if not jenis_barang:
        return gagal_dengan_form(
            "aset_list", "Jenis Barang (Kategori) wajib diisi.", "tambah_aset"
        )

    # Cari/buat kategori
    id_kategori = None
    if jenis_barang:
        kategori = Kategori.query.filter(db.func.lower(Kategori.nama) == jenis_barang.lower()).first()
        if not kategori:
            kategori = Kategori(nama=jenis_barang)
            db.session.add(kategori)
            db.session.flush()
        id_kategori = kategori.id

    # Upload foto
    foto_file = request.files.get("foto")
    foto = None
    foto_error = None
    if foto_file and foto_file.filename:
        foto, foto_error = save_upload(foto_file, prefix="aset_")

    foto_url_raw = request.form.get("foto_url", "").strip()
    foto_url = convert_gdrive_to_thumbnail(foto_url_raw) if foto_url_raw else None

    aset = Aset(
        kode_aset=kode_aset,
        nama=nama,
        area=area,
        fungsi=fungsi,
        merek=merek,
        serial_number=serial_number,
        spesifikasi=spesifikasi,
        tipe_aset=tipe_aset,
        volume=volume,
        satuan=satuan,
        status_aset=status_aset,
        gedung=gedung,
        ruangan=ruangan,
        lantai=lantai,
        foto=foto,
        foto_url=foto_url,
        link_qr=link_qr,
        tanggal_datang=tanggal_datang,
        keterangan=keterangan,
        id_kategori=id_kategori,
        total_kerusakan=0,
    )
    db.session.add(aset)
    db.session.flush()  # penting: dapatkan aset.id dari MySQL (AUTO_INCREMENT) sebelum dipakai di log aktivitas
    catat_aktivitas(
        aksi="CREATE",
        target_model="Aset",
        target_id=aset.id,
        deskripsi=f"Menambahkan aset baru: {aset.nama} ({aset.kode_aset})",
        data_baru=snapshot_aset(aset, kategori_nama=jenis_barang or None)
    )
    db.session.commit()
    
    if foto_error:
        flash(f"Aset berhasil ditambahkan, tetapi foto gagal diupload: {foto_error}", "warning")
    else:
        flash("Aset berhasil ditambahkan.", "success")
    return redirect(url_for("aset_list"))


@app.route("/aset/<int:aset_id>/edit", methods=["POST"])
@login_required
@role_required(ROLE_ADMIN)
def aset_edit(aset_id):
    aset = Aset.query.get_or_404(aset_id)

    # Jenis Barang (Kategori) WAJIB diisi -- validasi di awal sebelum field
    # lain diubah, supaya kalau gagal, objek `aset` di database belum
    # tersentuh sama sekali (belum ada db.session.commit()).
    jenis_barang = request.form.get("jenis_barang", "").strip()
    if not jenis_barang:
        return gagal_dengan_form(
            "aset_list", "Jenis Barang (Kategori) wajib diisi.", "edit_aset", id_target=aset.id
        )

    # +++ DEFINISIKAN data_lama SEBELUM diubah (snapshot lengkap semua field) +++
    kategori_lama_nama = aset.kategori_ref.nama if aset.kategori_ref else None
    data_lama = snapshot_aset(aset, kategori_nama=kategori_lama_nama)

    # Simpan status lama untuk logika total_kerusakan
    status_lama = aset.status_aset

    # ... update semua field ...
    aset.nama = request.form.get("nama", "").strip()
    aset.area = request.form.get("area", "").strip() or None
    aset.fungsi = request.form.get("fungsi", "").strip() or None
    aset.merek = request.form.get("merek", "").strip() or None
    aset.serial_number = request.form.get("serial_number", "").strip() or None
    aset.spesifikasi = request.form.get("spesifikasi", "").strip() or None
    aset.tipe_aset = request.form.get("tipe_aset", "OPEX").strip()
    aset.volume = request.form.get("volume", "").strip() or None
    aset.satuan = request.form.get("satuan", "").strip() or None
    aset.status_aset = request.form.get("status_aset", aset.status_aset)
    aset.gedung = request.form.get("gedung", "").strip()
    aset.ruangan = request.form.get("ruangan", "").strip()
    aset.lantai = request.form.get("lantai", "").strip() or None
    aset.keterangan = request.form.get("keterangan", "").strip() or None
    aset.link_qr = request.form.get("link_qr", "").strip() or None

    # Tanggal datang
    tanggal_datang_str = request.form.get("tanggal_datang", "").strip()
    aset.tanggal_datang = None  # default ke None
    if tanggal_datang_str:
        try:
            aset.tanggal_datang = datetime.strptime(tanggal_datang_str, "%Y-%m-%d").date()
        except ValueError:
            aset.tanggal_datang = None
    else:
        aset.tanggal_datang = None

    # Kategori
    id_kategori = None
    if jenis_barang:
        kategori = Kategori.query.filter(db.func.lower(Kategori.nama) == jenis_barang.lower()).first()
        if not kategori:
            kategori = Kategori(nama=jenis_barang)
            db.session.add(kategori)
            db.session.flush()
        id_kategori = kategori.id
    aset.id_kategori = id_kategori

    # Jika status berubah dari bukan Rusak menjadi Rusak
    if aset.status_aset == "Rusak" and status_lama != "Rusak":
        aset.total_kerusakan = (aset.total_kerusakan or 0) + 1
        # Catat histori rusak
        histori = HistoriAset(
            id_aset=aset.id,
            jenis_event="rusak",
            gedung=aset.gedung,
            lantai=aset.lantai,
            ruangan=aset.ruangan,
            id_tiket=None
        )
        db.session.add(histori)

    # Foto
    foto_url_raw = request.form.get("foto_url", "").strip()
    if foto_url_raw:
        aset.foto_url = convert_gdrive_to_thumbnail(foto_url_raw)
    elif foto_url_raw == "" and request.form.get("hapus_foto"):
        aset.foto_url = None

    foto_file = request.files.get("foto")
    foto_error = None
    if foto_file and foto_file.filename:
        foto, foto_error = save_upload(foto_file, prefix="aset_")
        if foto:
            aset.foto = foto

    # +++ DEFINISIKAN data_baru SETELAH perubahan (snapshot lengkap semua field) +++
    kategori_baru_nama = kategori.nama if jenis_barang else None
    data_baru = snapshot_aset(aset, kategori_nama=kategori_baru_nama)

    # Log aktivitas jika ada perubahan
    if data_lama != data_baru:
        catat_aktivitas(
            aksi="UPDATE",
            target_model="Aset",
            target_id=aset.id,
            deskripsi=f"Mengupdate aset: {aset.nama} ({aset.kode_aset})",
            data_lama=data_lama,
            data_baru=data_baru
        )

    db.session.commit()
    
    if foto_error:
        flash(f"Aset berhasil diperbarui, tetapi foto gagal diupload: {foto_error}", "warning")
    else:
        flash("Aset berhasil diperbarui.", "success")
    return redirect(url_for("aset_list"))


@app.route("/aset/<int:aset_id>/delete", methods=["POST"])
@login_required
@role_required(ROLE_ADMIN)
def aset_delete(aset_id):
    aset = Aset.query.get_or_404(aset_id)
    
    # Ambil alasan dari form
    reason = request.form.get("delete_reason", "").strip()
    
    # Catat aktivitas dengan alasan
    catat_aktivitas(
        aksi="DELETE",
        target_model="Aset",
        target_id=aset.id,
        deskripsi=f"Menghapus aset: {aset.nama} ({aset.kode_aset}) - Alasan: {reason or 'Tidak ada alasan'}",
        data_lama=snapshot_aset(aset),
        data_baru=None,
    )
    
    db.session.delete(aset)
    db.session.commit()
    flash(f"Aset berhasil dihapus." + (f" Alasan: {reason}" if reason else ""), "success")
    return redirect(url_for("aset_list"))

@app.route("/aset/<int:aset_id>/detail")
@login_required
def aset_detail(aset_id):
    aset = Aset.query.get_or_404(aset_id)
    histori = HistoriAset.query.filter_by(id_aset=aset_id).order_by(HistoriAset.tanggal.desc()).all()
    histori_data = []
    for h in histori:
        histori_data.append({
            "jenis": h.jenis_event,
            "gedung": h.gedung or "",
            "lantai": h.lantai or "",
            "ruangan": h.ruangan or "",
            "gedung_asal": h.gedung_asal or "",
            "lantai_asal": h.lantai_asal or "",
            "ruangan_asal": h.ruangan_asal or "",
            "tanggal": h.tanggal.strftime("%d-%m-%Y %H:%M"),
            "id_tiket": h.id_tiket
        })
    
    foto_display = None
    if aset.foto_url:
        foto_display = aset.foto_url
    elif aset.foto:
        foto_display = aset.foto
    
    data = {
        "id": aset.id,
        "kode_aset": aset.kode_aset,
        "nama": aset.nama,
        "area": aset.area or "",
        "fungsi": aset.fungsi or "",
        "merek": aset.merek or "",
        "serial_number": aset.serial_number or "",
        "spesifikasi": aset.spesifikasi or "",
        "tipe_aset": aset.tipe_aset,
        "volume": aset.volume or "",
        "satuan": aset.satuan or "",
        "status_aset": aset.status_aset,
        "gedung": aset.gedung,
        "lantai": aset.lantai or "",
        "ruangan": aset.ruangan,
        "kategori": aset.kategori_ref.nama if aset.kategori_ref else "-",
        "foto": foto_display,
        "total_kerusakan": aset.total_kerusakan or 0,
        "tanggal_datang": aset.tanggal_datang.strftime("%d-%m-%Y") if aset.tanggal_datang else "",
        "keterangan": aset.keterangan or "",
        "histori": histori_data
        # link_qr TIDAK dikirim (di-hide)
    }
    return jsonify(data)

@app.route("/aset/<int:aset_id>/histori")
@login_required
def aset_histori(aset_id):
    aset = Aset.query.get_or_404(aset_id)
    histori = HistoriAset.query.filter_by(id_aset=aset_id).order_by(HistoriAset.tanggal.desc()).all()
    data = []
    for h in histori:
        data.append({
            "jenis": h.jenis_event,
            "gedung": h.gedung or "",
            "lantai": h.lantai or "",
            "ruangan": h.ruangan or "",
            "tanggal": h.tanggal.strftime("%d-%m-%Y %H:%M"),
            "id_tiket": h.id_tiket
        })
    return jsonify(data)

@app.route("/aset/<int:aset_id>/pemindahan", methods=["GET"])
@login_required
@role_required(ROLE_ADMIN)
def aset_pemindahan_form(aset_id):
    """Halaman form Pemindahan khusus untuk 1 aset (dibuka lewat tombol
    'Pemindahan' di Data Aset). Lokasi Saat Ini diambil otomatis dari data
    aset, user hanya perlu mengisi Lokasi Tujuan."""
    aset = Aset.query.get_or_404(aset_id)
    lokasi_master = build_lokasi_master()
    return render_template(
        "aset/pemindahan_form.html",
        aset=aset,
        lokasi_master=lokasi_master,
    )


@app.route("/aset/<int:aset_id>/pemindahan", methods=["POST"])
@login_required
@role_required(ROLE_ADMIN)
def aset_pemindahan_submit(aset_id):
    """Proses submit form Pemindahan per-item: buat tiket Pemindahan (langsung
    Selesai) + histori aset, lalu update lokasi aset -- sama persis alurnya
    dengan tiket_create_pemindahan(), supaya otomatis muncul juga di menu
    Pemindahan Aset."""
    aset = Aset.query.get_or_404(aset_id)

    nama_pemohon = request.form.get("nama_pemohon", "").strip() or current_user.name
    tujuan_gedung = request.form.get("tujuan_gedung", "").strip()
    tujuan_lantai = request.form.get("tujuan_lantai", "").strip()
    tujuan_ruangan = request.form.get("tujuan_ruangan", "").strip()
    catatan = request.form.get("catatan", "").strip()

    if not tujuan_gedung or not tujuan_ruangan:
        flash("Lokasi Tujuan (Gedung & Ruangan) wajib diisi.", "danger")
        return redirect(url_for("aset_pemindahan_form", aset_id=aset.id))

    gedung_asal = aset.gedung
    lantai_asal = aset.lantai
    ruangan_asal = aset.ruangan

    lokasi_sama = (
        tujuan_gedung == (gedung_asal or "")
        and (tujuan_lantai or None) == (lantai_asal or None)
        and tujuan_ruangan == (ruangan_asal or "")
    )
    if lokasi_sama:
        flash("Lokasi Tujuan sama dengan lokasi aset saat ini.", "warning")
        return redirect(url_for("aset_pemindahan_form", aset_id=aset.id))

    tiket = Tiket(
        jenis_tiket="Pemindahan",
        nama_pemohon=nama_pemohon,
        gedung_asal=gedung_asal,
        lantai_asal=lantai_asal,
        ruangan_asal=ruangan_asal,
        gedung_tujuan=tujuan_gedung,
        lantai_tujuan=tujuan_lantai or None,
        ruangan_tujuan=tujuan_ruangan,
        catatan=catatan or None,
        created_by=current_user.id,
    )
    db.session.add(tiket)
    db.session.flush()

    db.session.add(HistoriAset(
        id_aset=aset.id,
        jenis_event="pindah",
        gedung=tujuan_gedung,
        lantai=tujuan_lantai or None,
        ruangan=tujuan_ruangan,
        gedung_asal=gedung_asal,
        lantai_asal=lantai_asal,
        ruangan_asal=ruangan_asal,
        id_tiket=tiket.id,
    ))

    db.session.add(LogStatus(
        id_tiket=tiket.id,
        status_lama=None,
        status_baru="Selesai",
        id_user_pengubah=current_user.id,
    ))

    db.session.add(TiketAset(id_tiket=tiket.id, id_aset=aset.id))

    data_lama_lokasi = {"gedung": gedung_asal, "lantai": lantai_asal, "ruangan": ruangan_asal}

    aset.gedung = tujuan_gedung
    aset.lantai = tujuan_lantai or None
    aset.ruangan = tujuan_ruangan
    aset.status_aset = "Baik"

    catat_aktivitas(
        aksi="MOVE",
        target_model="Aset",
        target_id=aset.id,
        deskripsi=f"Memindahkan aset {aset.nama} dari {gedung_asal or '-'} / {ruangan_asal or '-'} ke {tujuan_gedung} / {tujuan_ruangan}",
        data_lama=data_lama_lokasi,
        data_baru={"gedung": aset.gedung, "lantai": aset.lantai, "ruangan": aset.ruangan},
    )

    db.session.commit()

    flash(f"Aset {aset.nama} berhasil dipindahkan.", "success")
    return redirect(url_for("pemindahan_list"))


@app.route("/aset/delete-multiple", methods=["POST"])
@login_required
@role_required(ROLE_ADMIN)
def aset_delete_multiple():
    ids = request.form.getlist("ids[]")
    if not ids:
        flash("Tidak ada aset yang dipilih.", "danger")
        return redirect(url_for("aset_list"))
    
    try:
        ids = [int(id) for id in ids]
    except ValueError:
        flash("ID tidak valid.", "danger")
        return redirect(url_for("aset_list"))
    
    aset_list = Aset.query.filter(Aset.id.in_(ids)).all()
    if not aset_list:
        flash("Tidak ada aset yang ditemukan.", "danger")
        return redirect(url_for("aset_list"))

    # Ambil alasan dari form (sebelumnya tidak pernah dibaca, jadi selalu hilang)
    reason = request.form.get("delete_reason", "").strip()

    for aset in aset_list:
        catat_aktivitas(
            aksi="DELETE",
            target_model="Aset",
            target_id=aset.id,
            deskripsi=f"Menghapus aset via bulk: {aset.nama} ({aset.kode_aset}) - Alasan: {reason or 'Tidak ada alasan'}",
            data_lama=snapshot_aset(aset, kategori_nama=aset.kategori_ref.nama if aset.kategori_ref else None)
        )
        db.session.delete(aset)
    
    db.session.commit()
    flash(f"Berhasil menghapus {len(aset_list)} aset." + (f" Alasan: {reason}" if reason else ""), "success")
    return redirect(url_for("aset_list"))

# ---------------------------------------------------------------------------
# EXPORT / IMPORT ASET (Excel)
# ---------------------------------------------------------------------------
EXPORT_HEADERS = [
    "area", "kode_aset", "nama", "fungsi", "jenis_barang", 
    "merek", "serial_number", "spesifikasi", "tipe_aset", 
    "volume", "satuan", "status_aset", "gedung", "ruangan", 
    "lantai", "foto_url", "nama_project", "harga_perolehan", 
    "tanggal_bast", "evidence_bast", "mitra", "link_qr", 
    "tanggal_datang", "keterangan"
]


@app.route("/aset/export")
@login_required
@role_required(ROLE_ADMIN)
def aset_export():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data Aset"
    ws.append(EXPORT_HEADERS)

    for aset in Aset.query.order_by(Aset.id).all():
        ws.append([
            aset.area or "",
            aset.kode_aset,
            aset.nama,
            aset.fungsi or "",
            aset.kategori_ref.nama if aset.kategori_ref else "",
            aset.merek or "",
            aset.serial_number or "",
            aset.spesifikasi or "",
            aset.tipe_aset,  # <-- GANTI JENIS_ASET MENJADI TIPE_ASET
            aset.volume or "",
            aset.satuan or "",
            aset.status_aset,
            aset.gedung,
            aset.ruangan,
            aset.lantai or "",
            aset.foto_url or "",
            "",  # nama_project (tidak dipakai)
            "",  # harga_perolehan (tidak dipakai)
            "",  # tanggal_bast (tidak dipakai)
            "",  # evidence_bast (tidak dipakai)
            "",  # mitra (tidak dipakai)
            aset.link_qr or "",
            aset.tanggal_datang.strftime("%Y-%m-%d") if aset.tanggal_datang else "",
            aset.keterangan or "",
        ])

    for i in range(1, len(EXPORT_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    nama_file = f"data_aset_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=nama_file,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/aset/import", methods=["POST"])
@login_required
@role_required(ROLE_ADMIN)
def aset_import():
    file = request.files.get("file_import")
    if not file or file.filename == "":
        flash("Pilih file Excel (.xlsx) terlebih dahulu.", "danger")
        return redirect(url_for("aset_list"))
    if not file.filename.lower().endswith((".xlsx", ".xlsm")):
        flash("Format file harus .xlsx.", "danger")
        return redirect(url_for("aset_list"))

    try:
        wb = openpyxl.load_workbook(file.stream, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
    except Exception:
        flash("File Excel tidak valid atau rusak.", "danger")
        return redirect(url_for("aset_list"))

    ditambahkan, diperbarui, dilewati = 0, 0, 0
    tipe_kosong = 0  # baris yang tipe_aset-nya kosong/tidak valid di Excel
    error_baris = []

    for i, row in enumerate(rows, start=2):
        if not row or not any(row):
            continue

        cols = [str(c).strip() if c is not None else "" for c in row]
        cols += [""] * (24 - len(cols))
        cols = cols[:24]

        (area, kode_aset, nama, fungsi, jenis_barang, merek, serial_number,
         spesifikasi, tipe_aset, volume, satuan, status_aset, gedung, ruangan,
         lantai, foto_url, nama_project, harga_perolehan, tanggal_bast,
         evidence_bast, mitra, link_qr, tanggal_datang_str, keterangan) = cols

        kode_aset = kode_aset.strip()
        nama = nama.strip()
        if not kode_aset or not nama:
            dilewati += 1
            error_baris.append(f"Baris {i}: kode_aset atau nama kosong")
            continue

        # TAMBAHAN: simpan Area/Gedung/Lantai/Ruangan baris ini ke tabel
        # master (Area/Gedung/Lantai/Ruangan) supaya bisa dipakai sebagai
        # sumber dropdown di form Tambah/Edit Aset. Tidak ditampilkan
        # sebagai halaman tersendiri, cuma data pendukung dropdown.
        upsert_lokasi_master(area, gedung, lantai, ruangan)

        # +++ DEFINISIKAN DI SINI +++
        foto_url_thumbnail = convert_gdrive_to_thumbnail(foto_url) if foto_url else None

        # Cari/buat Kategori dari "jenis_barang"
        id_kategori = None
        if jenis_barang:
            kategori = Kategori.query.filter(db.func.lower(Kategori.nama) == jenis_barang.lower()).first()
            if not kategori:
                kategori = Kategori(nama=jenis_barang)
                db.session.add(kategori)
                db.session.flush()
            id_kategori = kategori.id

        # Parse tanggal datang
        tanggal_datang = None
        if tanggal_datang_str:
            try:
                tanggal_datang = datetime.strptime(tanggal_datang_str, "%Y-%m-%d").date()
            except ValueError:
                try:
                    tanggal_datang = datetime.strptime(tanggal_datang_str, "%d/%m/%Y").date()
                except ValueError:
                    pass

        # Status valid
        status_valid = status_aset if status_aset in ("Baik", "Rusak", "Dipindahkan") else "Baik"

        # PERBAIKAN: tipe_aset (CAPEX/OPEX) dari Excel.
        # tipe_valid = None berarti Excel TIDAK punya nilai CAPEX/OPEX yang valid
        # untuk baris ini. Sebelumnya nilai None ini langsung di-assign ke
        # aset.tipe_aset -- karena kolomnya nullable=False dengan default="OPEX"
        # di models.py, SQLAlchemy diam-diam mengganti None dengan "OPEX" setiap
        # kali membuat baris baru (perilaku default SQLAlchemy yang sering
        # mengecoh), sehingga SEMUA aset yang tipe_aset-nya kosong di Excel
        # otomatis jadi "OPEX" tanpa pemberitahuan apa pun ke user.
        tipe_valid = tipe_aset if tipe_aset in ("CAPEX", "OPEX") else None
        if not tipe_valid:
            tipe_kosong += 1

        # Cek apakah aset sudah ada
        aset = Aset.query.filter_by(kode_aset=kode_aset).first()

        if aset:
            # Simpan status lama
            status_lama = aset.status_aset
            
            # Update semua field
            aset.area = area or None
            aset.nama = nama
            aset.fungsi = fungsi or None
            aset.merek = merek or None
            aset.serial_number = serial_number or None
            aset.spesifikasi = spesifikasi or None
            # PERBAIKAN: hanya timpa tipe_aset kalau Excel punya nilai valid.
            # Kalau kosong, PERTAHANKAN nilai tipe_aset yang sudah ada di
            # database -- jangan ikut-ikutan di-reset ke default "OPEX".
            if tipe_valid:
                aset.tipe_aset = tipe_valid
            aset.volume = volume or None
            aset.satuan = satuan or None
            aset.status_aset = status_valid
            aset.gedung = gedung or "-"
            aset.ruangan = ruangan or "-"
            aset.lantai = lantai or None
            aset.foto_url = foto_url_thumbnail  # <-- PAKAI VARIABEL YANG SUDAH DIDEKLARASIKAN
            aset.link_qr = link_qr or None
            aset.tanggal_datang = tanggal_datang
            aset.keterangan = keterangan or None
            aset.id_kategori = id_kategori
            
            # Jika status berubah dari bukan Rusak menjadi Rusak
            if status_valid == "Rusak" and status_lama != "Rusak":
                aset.total_kerusakan = (aset.total_kerusakan or 0) + 1
            
            diperbarui += 1
        else:
            db.session.add(Aset(
                kode_aset=kode_aset,
                nama=nama,
                area=area or None,
                fungsi=fungsi or None,
                merek=merek or None,
                total_kerusakan=0,
                serial_number=serial_number or None,
                spesifikasi=spesifikasi or None,
                # PERBAIKAN: default "OPEX" ditulis eksplisit di sini (bukan
                # mengandalkan default kolom di models.py) supaya perilakunya
                # jelas dan gampang ditelusuri -- ini HANYA berlaku untuk aset
                # baru yang memang belum pernah ada tipe_aset-nya sama sekali.
                tipe_aset=tipe_valid or "OPEX",
                volume=volume or None,
                satuan=satuan or None,
                status_aset=status_valid,
                gedung=gedung or "-",
                ruangan=ruangan or "-",
                lantai=lantai or None,
                foto_url=foto_url_thumbnail,  # <-- PAKAI VARIABEL YANG SUDAH DIDEKLARASIKAN
                link_qr=link_qr or None,
                tanggal_datang=tanggal_datang,
                keterangan=keterangan or None,
                id_kategori=id_kategori,
            ))
            ditambahkan += 1

    db.session.commit()

    pesan = f"Import selesai: {ditambahkan} aset baru ditambahkan, {diperbarui} aset diperbarui."
    if dilewati:
        pesan += f" {dilewati} baris dilewati karena data tidak lengkap."
    flash(pesan, "warning" if dilewati else "success")
    if tipe_kosong:
        flash(
            f"Perhatian: {tipe_kosong} baris tidak punya nilai tipe_aset "
            f"(CAPEX/OPEX) yang valid di Excel. Untuk aset baru, tipe_aset "
            f"otomatis diset ke 'OPEX' (default); untuk aset yang sudah ada, "
            f"tipe_aset lama TETAP dipertahankan (tidak ditimpa). Silakan "
            f"lengkapi kolom tipe_aset di Excel jika perlu klasifikasi yang benar.",
            "warning",
        )
    if error_baris:
        flash(" | ".join(error_baris[:5]), "warning")

    return redirect(url_for("aset_list"))

@app.route("/aset/import-guide")
@login_required
@role_required(ROLE_ADMIN)
def aset_import_guide():
    """Halaman panduan import Excel dengan contoh data."""
    return render_template("aset/import_guide.html")


# ---------------------------------------------------------------------------
# KATEGORI & SUB KATEGORI
# ---------------------------------------------------------------------------
@app.route("/kategori")
@login_required
@role_required(ROLE_ADMIN)
def kategori_list():
    kategori_all = Kategori.query.order_by(Kategori.nama).all()

    # Hitung jumlah pemakaian di Peminjaman per kategori (jumlah Aset sudah
    # tersedia lewat k.aset_list, tidak perlu dihitung ulang di sini).
    jumlah_peminjaman_map = dict(
        db.session.query(Peminjaman.id_kategori, db.func.count(Peminjaman.id))
        .filter(Peminjaman.id_kategori.isnot(None))
        .group_by(Peminjaman.id_kategori)
        .all()
    )
    for k in kategori_all:
        k.jumlah_peminjaman = jumlah_peminjaman_map.get(k.id, 0)

    return render_template("kategori/list.html", kategori_all=kategori_all)


@app.route("/kategori/create", methods=["POST"])
@login_required
@role_required(ROLE_ADMIN)
def kategori_create():
    nama = (request.form.get("nama") or "").strip()
    if not nama:
        flash("Nama kategori tidak boleh kosong.", "danger")
        return redirect(url_for("kategori_list"))

    # Kolom `nama` bersifat unique di database -- cek dulu (case-insensitive)
    # supaya tidak crash (IntegrityError) kalau nama kategori sudah ada.
    sudah_ada = Kategori.query.filter(db.func.lower(Kategori.nama) == nama.lower()).first()
    if sudah_ada:
        flash(f"Kategori '{sudah_ada.nama}' sudah ada.", "danger")
        return redirect(url_for("kategori_list"))

    db.session.add(Kategori(nama=nama))
    db.session.commit()
    flash("Kategori ditambahkan.", "success")
    return redirect(url_for("kategori_list"))


@app.route("/kategori/<int:id>/delete", methods=["POST"])
@login_required
@role_required(ROLE_ADMIN)
def kategori_delete(id):
    """Hapus Kategori -- ditolak kalau masih dipakai oleh data Aset atau
    Peminjaman manapun, supaya tidak ada data yang jadi 'nyangkut'/rusak
    referensinya."""
    kategori = Kategori.query.get_or_404(id)

    jumlah_aset = Aset.query.filter_by(id_kategori=id).count()
    jumlah_peminjaman = Peminjaman.query.filter_by(id_kategori=id).count()

    if jumlah_aset > 0 or jumlah_peminjaman > 0:
        pemakai = []
        if jumlah_aset > 0:
            pemakai.append(f"{jumlah_aset} data Aset")
        if jumlah_peminjaman > 0:
            pemakai.append(f"{jumlah_peminjaman} data Peminjaman")
        flash(
            f"Kategori '{kategori.nama}' tidak bisa dihapus karena masih dipakai oleh "
            f"{' dan '.join(pemakai)}. Ubah/hapus data tersebut dulu, atau pindahkan "
            f"ke kategori lain.",
            "danger",
        )
        return redirect(url_for("kategori_list"))

    db.session.delete(kategori)
    db.session.commit()
    flash(f"Kategori '{kategori.nama}' dihapus.", "success")
    return redirect(url_for("kategori_list"))


@app.route("/kategori/delete-multiple", methods=["POST"])
@login_required
@role_required(ROLE_ADMIN)
def kategori_delete_multiple():
    """Hapus banyak Kategori sekaligus. Kategori yang masih dipakai Aset/
    Peminjaman otomatis DILEWATI (tidak ikut terhapus) supaya data lain
    tidak jadi rusak referensinya -- hasilnya dilaporkan lewat flash."""
    ids = request.form.getlist("ids[]")
    if not ids:
        flash("Tidak ada kategori yang dipilih.", "danger")
        return redirect(url_for("kategori_list"))

    try:
        ids = [int(i) for i in ids]
    except ValueError:
        flash("ID tidak valid.", "danger")
        return redirect(url_for("kategori_list"))

    kategori_list_dipilih = Kategori.query.filter(Kategori.id.in_(ids)).all()
    if not kategori_list_dipilih:
        flash("Tidak ada kategori yang ditemukan.", "danger")
        return redirect(url_for("kategori_list"))

    dihapus, gagal = [], []
    for kategori in kategori_list_dipilih:
        jumlah_aset = Aset.query.filter_by(id_kategori=kategori.id).count()
        jumlah_peminjaman = Peminjaman.query.filter_by(id_kategori=kategori.id).count()
        if jumlah_aset > 0 or jumlah_peminjaman > 0:
            gagal.append(kategori.nama)
        else:
            dihapus.append(kategori.nama)
            db.session.delete(kategori)

    db.session.commit()

    if dihapus:
        flash(f"{len(dihapus)} kategori berhasil dihapus: {', '.join(dihapus)}.", "success")
    if gagal:
        flash(
            f"{len(gagal)} kategori DILEWATI karena masih dipakai Aset/Peminjaman: {', '.join(gagal)}.",
            "danger",
        )
    return redirect(url_for("kategori_list"))


def find_kategori_saja(nama):
    """Cari Kategori yang SUDAH ADA berdasarkan nama (case-insensitive),
    TIDAK PERNAH membuat Kategori baru. Dipakai khusus di modul Peminjaman
    (form Tambah Peminjaman & import Excel) supaya daftar Kategori tidak
    ikut bertambah dari sana -- Kategori hanya boleh ditambahkan lewat
    modul Data Aset. Kalau namanya belum terdaftar sebagai Kategori,
    Peminjaman tetap menyimpan teksnya di kolom jenis_barang (legacy),
    hanya saja tidak ditautkan (id_kategori tetap kosong)."""
    nama = (nama or "").strip()
    if not nama:
        return None
    return Kategori.query.filter(db.func.lower(Kategori.nama) == nama.lower()).first()


# ---------------------------------------------------------------------------
# HISTORY (TIKET READ-ONLY)
# ---------------------------------------------------------------------------
@app.route("/history")
@login_required
def history_list():
    """Halaman history terpadu: tiket + aktivitas admin + maintenance."""
    filter_jenis = request.args.get("filter", "")
    
    events = []
    
    # === 1. TIKET (Pemindahan / Kerusakan) ===
    if filter_jenis in ["", "Tiket", "Pemindahan", "Kerusakan"]:
        tiket_query = Tiket.query
        if filter_jenis == "Pemindahan":
            tiket_query = tiket_query.filter_by(jenis_tiket="Pemindahan")
        elif filter_jenis == "Kerusakan":
            tiket_query = tiket_query.filter_by(jenis_tiket="Kerusakan")
        
        for t in tiket_query.order_by(Tiket.created_at.desc()).all():
            creator_name = "System"
            if t.created_by:
                creator = User.query.get(t.created_by)
                if creator:
                    creator_name = creator.name
            elif t.log_status:
                first_log = t.log_status[0]
                if first_log.user_pengubah:
                    creator_name = first_log.user_pengubah.name

            aset_list = ", ".join([ta.aset.nama for ta in t.aset_terkait[:3]])
            if len(t.aset_terkait) > 3:
                aset_list += f" dan {len(t.aset_terkait)-3} lainnya"

            events.append({
                "id": t.id,
                "waktu": t.created_at,
                "pelaku": creator_name,
                "jenis": "Tiket",
                "aksi": t.jenis_tiket,
                "detail": f"{t.nama_pemohon} - {aset_list or 'Tidak ada aset'}",
                "link": url_for("history_detail", tiket_id=t.id),
                "warna": "bg-rose-100 text-rose-700 border-rose-200" if t.jenis_tiket == "Kerusakan" else "bg-blue-100 text-blue-700 border-blue-200",
                "is_tiket": True,
                "is_maintenance": False,
                "is_aktivitas": False,
            })

    # === 1b. PEMINJAMAN ASET ===
    if filter_jenis in ["", "Peminjaman"]:
        for p in Peminjaman.query.order_by(Peminjaman.created_at.desc()).all():
            creator_name = p.user_creator.name if p.user_creator else "System"
            aset_list = ", ".join([pa.aset.nama for pa in p.aset_terkait[:3] if pa.aset])
            if len(p.aset_terkait) > 3:
                aset_list += f" dan {len(p.aset_terkait)-3} lainnya"

            events.append({
                "id": p.id,
                "waktu": p.created_at,
                "pelaku": creator_name,
                "jenis": "Peminjaman",
                "aksi": "Dikembalikan" if p.status == "Dikembalikan" else "Peminjaman",
                "detail": f"{p.nama_peminjam} - {aset_list or 'Tidak ada aset'}",
                "link": url_for("peminjaman_detail", id=p.id),
                "warna": "bg-emerald-100 text-emerald-700 border-emerald-200" if p.status == "Dikembalikan" else "bg-amber-100 text-amber-700 border-amber-200",
                "is_tiket": False,
                "is_maintenance": False,
                "is_aktivitas": False,
            })

    # === 2. AKTIVITAS ADMIN (hanya untuk role admin) ===
    if current_user.role == ROLE_ADMIN and filter_jenis in ["", "Aktivitas"]:
        for a in AktivitasLog.query.order_by(AktivitasLog.created_at.desc()).all():
            user = User.query.get(a.id_user)
            pelaku = user.name if user else "Unknown"

            label_aksi = label_aktivitas(a)

            events.append({
                "id": a.id,
                "waktu": a.created_at,
                "pelaku": pelaku,
                "jenis": "Aktivitas",
                "aksi": label_aksi,
                "detail": a.deskripsi or f"{label_aksi} ID {a.target_id}",
                "link": url_for("aktivitas_detail", log_id=a.id),
                "warna": "bg-indigo-100 text-indigo-700 border-indigo-200" if a.aksi == "CREATE" else "bg-amber-100 text-amber-700 border-amber-200" if a.aksi == "UPDATE" else "bg-rose-100 text-rose-700 border-rose-200" if a.aksi == "DELETE" else "bg-emerald-100 text-emerald-700 border-emerald-200",
                "is_tiket": False,
                "is_maintenance": False,
                "is_aktivitas": True,
                "data_lama": a.data_lama,
                "data_baru": a.data_baru
            })

    # === 3. MAINTENANCE ===
    if filter_jenis in ["", "Maintenance"]:
        for m in Maintenance.query.order_by(Maintenance.created_at.desc()).all():
            aset = Aset.query.get(m.id_aset)
            pelaku = m.user.name if m.user else "System"
            events.append({
                "id": m.id,
                "waktu": m.created_at,
                "pelaku": pelaku,
                "jenis": "Maintenance",
                "aksi": "Jadwal Maintenance",
                "detail": f"{aset.nama if aset else '-'} - {m.judul}",
                "link": url_for("maintenance_detail", id=m.id),
                "warna": "bg-emerald-100 text-emerald-700 border-emerald-200",
                "is_tiket": False,
                "is_maintenance": True,
                "is_aktivitas": False,
            })

    # Urutkan berdasarkan waktu terbaru
    events.sort(key=lambda x: x["waktu"], reverse=True)

    # Pagination
    page = request.args.get("page", 1, type=int)
    per_page = 10
    total = len(events)
    start = (page - 1) * per_page
    end = start + per_page
    events_page = events[start:end]
    total_pages = (total + per_page - 1) // per_page

    class PaginationDummy:
        def __init__(self, items, page, per_page, total, total_pages):
            self.items = items
            self.page = page
            self.per_page = per_page
            self.total = total
            self.pages = total_pages
            self.has_prev = page > 1
            self.has_next = page < total_pages
            self.prev_num = page - 1 if page > 1 else None
            self.next_num = page + 1 if page < total_pages else None

    pagination = PaginationDummy(events_page, page, per_page, total, total_pages)
    daftar_history = events_page

    return render_template(
        "history/list.html",
        daftar_history=daftar_history,
        pagination=pagination,
        filter_terpilih=filter_jenis,
    )

@app.route("/aktivitas/<int:log_id>")
@login_required
@role_required(ROLE_ADMIN)
def aktivitas_detail(log_id):
    """Detail aktivitas admin (tambah/edit/hapus aset/peminjaman/maintenance)."""
    log = AktivitasLog.query.get_or_404(log_id)

    # Label aksi disesuaikan per target_model & jenis perubahan (mis. "Perpanjangan
    # Peminjaman", "Upload Foto Dokumentasi") -- fungsi yang sama dipakai di History List
    # supaya labelnya selalu konsisten di kedua halaman.
    label_aksi = label_aktivitas(log)

    user = User.query.get(log.id_user)
    pelaku = user.name if user else "Unknown"

    field_diff = []
    foto_aktivitas = None       # foto baru/sekarang (upload foto maintenance)
    foto_lama_aktivitas = None  # foto yang digantikan/dihapus (kalau ada)
    foto_label = None
    peminjaman_current = None   # data peminjaman terkini (live), untuk ringkasan lengkap
    maintenance_current = None  # data maintenance terkini (live), untuk ringkasan lengkap

    if log.target_model == "Aset" and log.aksi in ("CREATE", "UPDATE", "DELETE"):
        field_diff = build_field_diff(log.data_lama, log.data_baru)

    # Detail data Peminjaman (CREATE/UPDATE/DELETE) -- perpanjangan, pengembalian,
    # upload evidence, dll semuanya lewat AktivitasLog target_model="Peminjaman"
    # dengan data_lama/data_baru, jadi bisa dipakai ulang build_field_diff yang sama.
    if log.target_model == "Peminjaman" and log.aksi in ("CREATE", "UPDATE", "DELETE"):
        field_diff = build_field_diff(log.data_lama, log.data_baru, FIELD_LABELS_PEMINJAMAN)
        # Ambil data peminjaman yang masih hidup supaya halaman ini menampilkan
        # ringkasan LENGKAP (siapa, unit, barang, status terkini, dst) -- bukan cuma
        # field yang berubah di aktivitas ini -- persis seperti Detail Data Aset.
        if log.aksi == "UPDATE":
            peminjaman_current = Peminjaman.query.get(log.target_id)

    # Detail data Maintenance: bedakan aktivitas foto dokumentasi (before/progress/after)
    # dari edit jadwal biasa (judul/tipe/status/dll), karena bentuk datanya beda.
    if log.target_model == "Maintenance":
        data_baru = log.data_baru or {}
        data_lama = log.data_lama or {}
        is_foto_activity = ("foto" in data_baru) or ("foto" in data_lama)

        if log.aksi == "UPDATE" and is_foto_activity:
            if data_baru.get("foto"):
                foto_aktivitas = data_baru.get("foto")
                foto_label = data_baru.get("label", "Foto")
            if data_lama.get("foto"):
                foto_lama_aktivitas = data_lama.get("foto")
                foto_label = foto_label or data_lama.get("label", "Foto")
        elif log.aksi in ("CREATE", "UPDATE", "DELETE"):
            field_diff = build_field_diff(log.data_lama, log.data_baru, FIELD_LABELS_MAINTENANCE)

        if log.aksi == "UPDATE":
            maintenance_current = Maintenance.query.get(log.target_id)

    return render_template(
        "history/aktivitas_detail.html",
        log=log,
        label_aksi=label_aksi,
        pelaku=pelaku,
        field_diff=field_diff,
        foto_aktivitas=foto_aktivitas,
        foto_lama_aktivitas=foto_lama_aktivitas,
        foto_label=foto_label,
        peminjaman_current=peminjaman_current,
        maintenance_current=maintenance_current,
    )

@app.route("/history/<int:tiket_id>")
@login_required
def history_detail(tiket_id):
    """Detail history tiket (read-only)"""
    tiket = Tiket.query.get_or_404(tiket_id)
    return render_template("history/detail.html", tiket=tiket)


def build_nomor_urut_tiket(jenis_tiket):
    """Bangun mapping {id_tiket: nomor_urut} KHUSUS untuk 1 jenis tiket
    (\"Pemindahan\" atau \"Kerusakan\"), diurutkan berdasarkan waktu dibuat
    (paling lama = nomor 1, dst).

    Tabel `tiket` dipakai bersama oleh modul Pemindahan & Kerusakan, jadi
    `Tiket.id` (primary key) itu satu deret angka yang sama-sama dipakai
    kedua modul -- makanya kalau ditampilkan apa adanya, nomornya kelihatan
    "nyambung" lintas modul (mis. Pemindahan #1, lalu tiket Kerusakan
    berikutnya jadi #2, padahal itu Kerusakan pertama). Fungsi ini menghitung
    ulang nomor urut yang independen per jenis tiket supaya masing-masing
    modul mulai dari #1 sendiri-sendiri."""
    ids_urut = [
        row[0] for row in db.session.query(Tiket.id)
        .filter_by(jenis_tiket=jenis_tiket)
        .order_by(Tiket.created_at.asc(), Tiket.id.asc())
        .all()
    ]
    return {tid: idx + 1 for idx, tid in enumerate(ids_urut)}


def hitung_nomor_urut_tiket(tiket):
    """Versi hemat query dari build_nomor_urut_tiket() untuk 1 tiket saja
    (dipakai di halaman Detail Pemindahan/Kerusakan) -- cukup hitung berapa
    banyak tiket dengan jenis yang sama yang dibuat sebelum (atau bersamaan
    dengan, dengan id lebih kecil/sama) tiket ini."""
    return Tiket.query.filter(
        Tiket.jenis_tiket == tiket.jenis_tiket
    ).filter(
        db.or_(
            Tiket.created_at < tiket.created_at,
            db.and_(Tiket.created_at == tiket.created_at, Tiket.id <= tiket.id),
        )
    ).count()


@app.route("/pemindahan")
@login_required
@role_required(ROLE_ADMIN)
def pemindahan_list():
    """Halaman Pemindahan Aset: form pemindahan + daftar riwayatnya."""
    status = request.args.get("status", "").strip()

    query = Tiket.query.filter_by(jenis_tiket="Pemindahan")
    if status:
        query = query.filter(Tiket.status_tiket == status)

    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 10, type=int)
    pagination = query.order_by(Tiket.created_at.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )
    daftar_tiket = pagination.items

    gedung_all = [
        g[0] for g in db.session.query(Aset.gedung).distinct().order_by(Aset.gedung).all() if g[0]
    ]

    nomor_urut_map = build_nomor_urut_tiket("Pemindahan")

    return render_template(
        "pemindahan/list.html",
        daftar_tiket=daftar_tiket,
        pagination=pagination,
        gedung_all=gedung_all,
        nomor_urut_map=nomor_urut_map,
    )


@app.route("/pemindahan/<int:tiket_id>")
@login_required
@role_required(ROLE_ADMIN)
def pemindahan_detail(tiket_id):
    """Detail pemindahan aset."""
    tiket = Tiket.query.filter_by(id=tiket_id, jenis_tiket="Pemindahan").first_or_404()
    nomor_urut = hitung_nomor_urut_tiket(tiket)
    return render_template("pemindahan/detail.html", tiket=tiket, nomor_urut=nomor_urut)


@app.route("/pemindahan/export")
@login_required
@role_required(ROLE_ADMIN)
def pemindahan_export():
    """Export seluruh data Pemindahan Aset ke Excel."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data Pemindahan"
    ws.append([
        "No", "Nama Pemohon", "Aset Dipindahkan",
        "Gedung Asal", "Lantai Asal", "Ruangan Asal",
        "Gedung Tujuan", "Lantai Tujuan", "Ruangan Tujuan",
        "Catatan", "Status", "Dibuat Oleh", "Tanggal Dibuat",
    ])

    daftar = Tiket.query.filter_by(jenis_tiket="Pemindahan").order_by(Tiket.created_at.desc()).all()
    for no, t in enumerate(daftar, start=1):
        nama_aset = [f"{ta.aset.kode_aset} - {ta.aset.nama}" for ta in t.aset_terkait if ta.aset]
        ws.append([
            no,
            t.nama_pemohon,
            "; ".join(nama_aset) if nama_aset else "-",
            t.gedung_asal or "",
            t.lantai_asal or "",
            t.ruangan_asal or "",
            t.gedung_tujuan or "",
            t.lantai_tujuan or "",
            t.ruangan_tujuan or "",
            t.catatan or "",
            t.status_tiket or "-",
            t.user_creator.name if t.user_creator else "-",
            t.created_at.strftime("%Y-%m-%d %H:%M") if t.created_at else "",
        ])

    for i in range(1, 14):
        ws.column_dimensions[get_column_letter(i)].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    nama_file = f"data_pemindahan_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=nama_file,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/kerusakan")
@login_required
@role_required(ROLE_ADMIN)
def kerusakan_list():
    """Halaman Kerusakan Aset: form lapor kerusakan + daftar riwayatnya."""
    status = request.args.get("status", "").strip()

    query = Tiket.query.filter_by(jenis_tiket="Kerusakan")
    if status:
        query = query.filter(Tiket.status_tiket == status)

    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 10, type=int)
    pagination = query.order_by(Tiket.created_at.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )
    daftar_tiket = pagination.items

    gedung_all = [
        g[0] for g in db.session.query(Aset.gedung).distinct().order_by(Aset.gedung).all() if g[0]
    ]

    nomor_urut_map = build_nomor_urut_tiket("Kerusakan")

    return render_template(
        "kerusakan/list.html",
        daftar_tiket=daftar_tiket,
        pagination=pagination,
        gedung_all=gedung_all,
        nomor_urut_map=nomor_urut_map,
    )


@app.route("/kerusakan/<int:tiket_id>")
@login_required
@role_required(ROLE_ADMIN)
def kerusakan_detail(tiket_id):
    """Detail kerusakan aset."""
    tiket = Tiket.query.filter_by(id=tiket_id, jenis_tiket="Kerusakan").first_or_404()
    nomor_urut = hitung_nomor_urut_tiket(tiket)
    return render_template("kerusakan/detail.html", tiket=tiket, nomor_urut=nomor_urut)


@app.route("/kerusakan/export")
@login_required
@role_required(ROLE_ADMIN)
def kerusakan_export():
    """Export seluruh data Kerusakan Aset ke Excel."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data Kerusakan"
    ws.append([
        "No", "Nama Pemohon", "Aset Rusak",
        "Gedung", "Lantai", "Ruangan",
        "Catatan", "Status", "Dibuat Oleh", "Tanggal Dibuat",
    ])

    daftar = Tiket.query.filter_by(jenis_tiket="Kerusakan").order_by(Tiket.created_at.desc()).all()
    for no, t in enumerate(daftar, start=1):
        nama_aset = [f"{ta.aset.kode_aset} - {ta.aset.nama}" for ta in t.aset_terkait if ta.aset]
        ws.append([
            no,
            t.nama_pemohon,
            "; ".join(nama_aset) if nama_aset else "-",
            t.gedung_asal or "",
            t.lantai_asal or "",
            t.ruangan_asal or "",
            t.catatan or "",
            t.status_tiket or "-",
            t.user_creator.name if t.user_creator else "-",
            t.created_at.strftime("%Y-%m-%d %H:%M") if t.created_at else "",
        ])

    for i in range(1, 11):
        ws.column_dimensions[get_column_letter(i)].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    nama_file = f"data_kerusakan_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=nama_file,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/tiket/create/pemindahan", methods=["POST"])
@login_required
@role_required(ROLE_ADMIN)
def tiket_create_pemindahan():
    """Buat tiket pemindahan (langsung selesai)."""
    aset_ids = request.form.getlist("aset_ids[]")
    if not aset_ids:
        flash("Pilih minimal 1 aset.", "danger")
        return redirect(url_for("pemindahan_list"))

    gedung_asal = request.form.get("gedung_asal", "").strip()
    lantai_asal = request.form.get("lantai_asal", "").strip()
    ruangan_asal = request.form.get("ruangan_asal", "").strip()
    gedung_tujuan = request.form.get("gedung_tujuan", "").strip()
    lantai_tujuan = request.form.get("lantai_tujuan", "").strip()
    ruangan_tujuan = request.form.get("ruangan_tujuan", "").strip()
    nama_pemohon = request.form.get("nama_pemohon", "").strip()
    catatan = request.form.get("catatan", "").strip()

    foto, foto_error = save_upload(request.files.get("foto"), prefix="tiket_")

    tiket = Tiket(
        jenis_tiket="Pemindahan",
        nama_pemohon=nama_pemohon,
        gedung_asal=gedung_asal,
        lantai_asal=lantai_asal,
        ruangan_asal=ruangan_asal,
        gedung_tujuan=gedung_tujuan,
        lantai_tujuan=lantai_tujuan,
        ruangan_tujuan=ruangan_tujuan,
        catatan=catatan,
        foto=foto,
        created_by=current_user.id,
    )
    db.session.add(tiket)
    db.session.flush()

    for aid in aset_ids:
        aset = db.session.get(Aset, int(aid))
        if aset:
            # Simpan data lama
            data_lama = {
                "gedung": aset.gedung,
                "lantai": aset.lantai,
                "ruangan": aset.ruangan
            }

            # Catat histori pindah
            histori = HistoriAset(
                id_aset=aset.id,
                jenis_event="pindah",
                gedung=gedung_tujuan,
                lantai=lantai_tujuan,
                ruangan=ruangan_tujuan,
                gedung_asal=aset.gedung,
                lantai_asal=aset.lantai,
                ruangan_asal=aset.ruangan,
                id_tiket=tiket.id
            )
            db.session.add(histori)

            # Update lokasi aset
            aset.gedung = gedung_tujuan
            aset.lantai = lantai_tujuan or None
            aset.ruangan = ruangan_tujuan
            aset.status_aset = "Baik"

            # Catat log status (untuk tiket)
            db.session.add(LogStatus(
                id_tiket=tiket.id,
                status_lama=None,
                status_baru="Selesai",
                id_user_pengubah=current_user.id
            ))

            # Catat aktivitas admin
            catat_aktivitas(
                aksi="MOVE",
                target_model="Aset",
                target_id=aset.id,
                deskripsi=f"Memindahkan aset {aset.nama} dari {data_lama['gedung']} / {data_lama['ruangan']} ke {gedung_tujuan} / {ruangan_tujuan}",
                data_lama=data_lama,
                data_baru={
                    "gedung": aset.gedung,
                    "lantai": aset.lantai,
                    "ruangan": aset.ruangan
                }
            )

            db.session.add(TiketAset(id_tiket=tiket.id, id_aset=aset.id))

    db.session.commit()

    if foto_error:
        flash(f"Pemindahan berhasil, tetapi foto gagal diupload: {foto_error}", "warning")
    else:
        flash(f"Pemindahan berhasil. {len(aset_ids)} aset dipindahkan.", "success")
    return redirect(url_for("pemindahan_list"))

@app.route("/tiket/create/kerusakan", methods=["POST"])
@login_required
@role_required(ROLE_ADMIN)
def tiket_create_kerusakan():
    """Buat tiket kerusakan (hanya kerusakan, langsung selesai)."""
    aset_ids = request.form.getlist("aset_ids[]")
    if not aset_ids:
        flash("Pilih minimal 1 aset.", "danger")
        return redirect(url_for("kerusakan_list"))

    gedung_asal = request.form.get("gedung_asal", "").strip()
    lantai_asal = request.form.get("lantai_asal", "").strip()
    ruangan_asal = request.form.get("ruangan_asal", "").strip()
    nama_pemohon = request.form.get("nama_pemohon", "").strip()
    catatan = request.form.get("catatan", "").strip()

    foto = save_upload(request.files.get("foto"), prefix="tiket_")[0]

    tiket = Tiket(
        jenis_tiket="Kerusakan",
        nama_pemohon=nama_pemohon,
        gedung_asal=gedung_asal,
        lantai_asal=lantai_asal,
        ruangan_asal=ruangan_asal,
        catatan=catatan,
        foto=foto,
        created_by=current_user.id,
    )
    db.session.add(tiket)
    db.session.flush()

    for aid in aset_ids:
        aset = db.session.get(Aset, int(aid))
        if aset:
            # Update status aset menjadi Rusak
            aset.status_aset = "Rusak"
            aset.total_kerusakan = (aset.total_kerusakan or 0) + 1
            
            # Catat histori rusak
            histori = HistoriAset(
                id_aset=aset.id,
                jenis_event="rusak",
                gedung=aset.gedung,
                lantai=aset.lantai,
                ruangan=aset.ruangan,
                id_tiket=tiket.id
            )
            db.session.add(histori)
            
            db.session.add(TiketAset(id_tiket=tiket.id, id_aset=aset.id))

    # Catat status "Selesai" lewat LogStatus (bukan kolom status_tiket --
    # lihat models.py: Tiket.status_tiket sekarang dihitung dari sini)
    catat_log(tiket, None, "Selesai")

    db.session.commit()
    flash(f"Laporan kerusakan berhasil dibuat. {len(aset_ids)} aset ditandai rusak.", "success")
    return redirect(url_for("kerusakan_list"))

@app.route("/peminjaman")
@login_required
def peminjaman_list():
    """Daftar peminjaman aset (gaya BA Transfer: Nama, Unit, Lokasi Kerja,
    Jenis Barang, Tanggal Awal/Akhir, Keterangan, Evidence Lampiran)."""
    search = request.args.get("search", "").strip()
    status = request.args.get("status", "").strip()
    jenis_transaksi = request.args.get("jenis_transaksi", "").strip()
    unit_filter = request.args.get("unit", "").strip()

    today = datetime.now(WIB).date()

    query = Peminjaman.query

    if status == "Terlambat":
        # "Terlambat" bukan nilai kolom status di database -- ini status
        # turunan (Dipinjam + lewat tanggal rencana kembali), jadi difilter
        # pakai kondisi tanggal, bukan Peminjaman.status langsung.
        query = query.filter(
            Peminjaman.status == "Dipinjam",
            Peminjaman.tanggal_rencana_kembali.isnot(None),
            Peminjaman.tanggal_rencana_kembali < today,
        )
    elif status:
        query = query.filter(Peminjaman.status == status)

    if jenis_transaksi:
        query = query.filter(Peminjaman.jenis_transaksi == jenis_transaksi)

    if unit_filter:
        query = query.filter(Peminjaman.unit == unit_filter)

    if search:
        query = query.join(PeminjamanAset, isouter=True).join(Aset, isouter=True).filter(
            db.or_(
                Peminjaman.nama_peminjam.ilike(f"%{search}%"),
                Peminjaman.unit.ilike(f"%{search}%"),
                Aset.nama.ilike(f"%{search}%"),
                Aset.kode_aset.ilike(f"%{search}%"),
            )
        ).distinct()

    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 10, type=int)
    if per_page not in [10, 25, 50, 100]:
        per_page = 10
    pagination = query.order_by(Peminjaman.tanggal_pinjam.desc(), Peminjaman.id.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )
    daftar_peminjaman = pagination.items

    total = Peminjaman.query.count()
    total_dipinjam = Peminjaman.query.filter_by(status="Dipinjam").count()
    total_dikembalikan = Peminjaman.query.filter_by(status="Dikembalikan").count()
    total_terlambat = Peminjaman.query.filter(
        Peminjaman.status == "Dipinjam",
        Peminjaman.tanggal_rencana_kembali.isnot(None),
        Peminjaman.tanggal_rencana_kembali < today,
    ).count()

    gedung_all = (
        db.session.query(Aset.area, Aset.gedung)
        .filter(Aset.gedung.isnot(None), Aset.gedung != "")
        .distinct()
        .order_by(Aset.area, Aset.gedung)
        .all()
    )
    gedung_all_formatted = [
        {"value": g.gedung, "label": f"{format_area_label(g.area)} - {g.gedung}" if g.area else g.gedung}
        for g in gedung_all
    ]

    kategori_all = Kategori.query.order_by(Kategori.nama).all()  # tetap dikirim untuk dropdown "Jenis Barang" di modal Tambah Peminjaman
    # Pilihan filter Unit diambil dari nilai unik kolom Peminjaman.unit
    # (teks bebas) -- TIDAK ada tabel master Unit terpisah di database.
    unit_all = [
        r[0] for r in db.session.query(Peminjaman.unit)
        .filter(Peminjaman.unit.isnot(None), Peminjaman.unit != "")
        .distinct()
        .order_by(Peminjaman.unit)
        .all()
    ]

    # Kalau sebelumnya ada submit form Tambah Peminjaman yang gagal validasi,
    # ambil (dan hapus) data yang sempat disimpan supaya modal bisa otomatis
    # terbuka lagi dengan field-field yang sudah terisi. Lihat gagal_dengan_form().
    repop = session.pop("form_repopulate", None)

    return render_template(
        "peminjaman/list.html",
        daftar_peminjaman=daftar_peminjaman,
        pagination=pagination,
        search=search,
        status_terpilih=status,
        jenis_transaksi_terpilih=jenis_transaksi,
        jenis_transaksi_options=JENIS_TRANSAKSI_OPTIONS,
        unit_terpilih=unit_filter,
        unit_all=unit_all,
        kategori_all=kategori_all,
        gedung_all=gedung_all_formatted,
        today=today,
        total=total,
        total_dipinjam=total_dipinjam,
        total_dikembalikan=total_dikembalikan,
        total_terlambat=total_terlambat,
        repop=repop,
    )


@app.route("/peminjaman/export")
@login_required
@role_required(ROLE_ADMIN)
def peminjaman_export():
    """Export seluruh data Peminjaman ke Excel (kolom rapi, 1 baris = 1
    data peminjaman -- bukan gabungan histori event, biar tidak bercampur
    dengan modul lain)."""
    today = datetime.now(WIB).date()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data Peminjaman"
    ws.append([
        "No", "Nama Peminjam", "Unit", "Lokasi Kerja",
        "Jenis Barang (Kategori)", "Jenis Transaksi", "Aset/Barang Dipinjam",
        "Tanggal Pinjam", "Rencana Kembali", "Tanggal Dikembalikan",
        "Status", "Status Perpanjangan", "Keterangan",
        "Dibuat Oleh", "Tanggal Dibuat",
    ])

    daftar = Peminjaman.query.order_by(Peminjaman.tanggal_pinjam.desc(), Peminjaman.id.desc()).all()
    for no, p in enumerate(daftar, start=1):
        if p.status == "Dipinjam" and p.tanggal_rencana_kembali and p.tanggal_rencana_kembali < today:
            status_tampil = "Terlambat"
        else:
            status_tampil = p.status

        nama_barang = [f"{pa.aset.kode_aset} - {pa.aset.nama}" for pa in p.aset_terkait if pa.aset]
        barang_tampil = "; ".join(nama_barang) if nama_barang else "-"

        ws.append([
            no,
            p.nama_peminjam,
            p.unit or "",
            p.lokasi_kerja or "",
            p.kategori_ref.nama if p.kategori_ref else (p.jenis_barang or ""),
            p.jenis_transaksi or "",
            barang_tampil,
            p.tanggal_pinjam.strftime("%Y-%m-%d") if p.tanggal_pinjam else "",
            p.tanggal_rencana_kembali.strftime("%Y-%m-%d") if p.tanggal_rencana_kembali else "",
            p.tanggal_dikembalikan.strftime("%Y-%m-%d") if p.tanggal_dikembalikan else "",
            status_tampil,
            p.status_perpanjangan or "",
            p.keterangan or "",
            p.user_creator.name if p.user_creator else "-",
            p.created_at.strftime("%Y-%m-%d %H:%M") if p.created_at else "",
        ])

    for i in range(1, 16):
        ws.column_dimensions[get_column_letter(i)].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    nama_file = f"data_peminjaman_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=nama_file,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/peminjaman/create", methods=["POST"])
@login_required
@role_required(ROLE_ADMIN)
def peminjaman_create():
    """Buat data peminjaman aset baru.

    Catatan revisi: pemilihan Gedung/Lantai/Ruangan/Aset yang dipinjam
    SUDAH TIDAK WAJIB DIISI (bahkan sudah dihapus dari form Tambah
    Peminjaman) -- peminjaman sekarang boleh dicatat tanpa menautkan ke
    barang/aset spesifik manapun. Kalau suatu saat field `aset_ids[]` tetap
    dikirim (mis. dari flow lain), tetap diproses seperti biasa; kalau
    kosong, peminjaman tetap tersimpan tanpa data barang.
    """
    aset_ids = request.form.getlist("aset_ids[]")

    nama_peminjam = request.form.get("nama_peminjam", "").strip()
    unit = request.form.get("unit", "").strip()
    lokasi_kerja = request.form.get("lokasi_kerja", "").strip()
    keterangan = request.form.get("keterangan", "").strip()
    tanggal_pinjam_str = request.form.get("tanggal_pinjam")
    tanggal_rencana_str = request.form.get("tanggal_rencana_kembali")
    jenis_transaksi = request.form.get("jenis_transaksi", "").strip() or "Peminjaman"

    # Jenis Barang (Kategori) -- dropdown Kategori (sama seperti di Data Aset),
    # HANYA boleh memilih Kategori yang sudah ada. Peminjaman tidak pernah
    # menambah Kategori baru -- itu khusus wewenang modul Data Aset.
    # SEKARANG WAJIB DIISI (sebelumnya opsional).
    id_kategori_form = request.form.get("id_kategori", "").strip()
    id_kategori = int(id_kategori_form) if id_kategori_form.isdigit() else None

    if not id_kategori:
        return gagal_dengan_form(
            "peminjaman_list", "Jenis Barang (Kategori) wajib dipilih.", "tambah_peminjaman"
        )

    if not nama_peminjam or not tanggal_pinjam_str:
        return gagal_dengan_form(
            "peminjaman_list", "Nama peminjam dan tanggal pinjam wajib diisi.", "tambah_peminjaman"
        )

    try:
        tanggal_pinjam = datetime.strptime(tanggal_pinjam_str, "%Y-%m-%d").date()
    except ValueError:
        return gagal_dengan_form(
            "peminjaman_list", "Format tanggal pinjam tidak valid.", "tambah_peminjaman"
        )

    tanggal_rencana_kembali = None
    if tanggal_rencana_str:
        try:
            tanggal_rencana_kembali = datetime.strptime(tanggal_rencana_str, "%Y-%m-%d").date()
        except ValueError:
            return gagal_dengan_form(
                "peminjaman_list", "Format tanggal rencana kembali tidak valid.", "tambah_peminjaman"
            )

    # VALIDASI URUTAN TANGGAL: rencana kembali tidak boleh lebih awal dari
    # tanggal pinjam (mis. pinjam 20 Januari tapi rencana kembali 12
    # Januari seharusnya ditolak).
    if tanggal_rencana_kembali and tanggal_rencana_kembali < tanggal_pinjam:
        return gagal_dengan_form(
            "peminjaman_list",
            "Tanggal rencana kembali tidak boleh lebih awal dari tanggal pinjam.",
            "tambah_peminjaman",
        )

    # Evidence Lampiran (Berita Acara) SEKARANG WAJIB dan HARUS berupa PDF.
    evidence, evidence_error = save_evidence_pdf(request.files.get("evidence"))
    if evidence_error:
        return gagal_dengan_form("peminjaman_list", evidence_error, "tambah_peminjaman")

    peminjaman = Peminjaman(
        nama_peminjam=nama_peminjam,
        unit=unit or None,
        lokasi_kerja=lokasi_kerja or None,
        id_kategori=id_kategori,
        jenis_transaksi=jenis_transaksi,
        tanggal_pinjam=tanggal_pinjam,
        tanggal_rencana_kembali=tanggal_rencana_kembali,
        status="Dipinjam",
        keterangan=keterangan or None,
        evidence=evidence,
        created_by=current_user.id,
    )
    db.session.add(peminjaman)
    db.session.flush()

    nama_aset_list = []
    for aid in aset_ids:
        aset = db.session.get(Aset, int(aid))
        if aset:
            db.session.add(PeminjamanAset(id_peminjaman=peminjaman.id, id_aset=aset.id))
            db.session.add(HistoriAset(
                id_aset=aset.id,
                jenis_event="pinjam",
                gedung=aset.gedung,
                lantai=aset.lantai,
                ruangan=aset.ruangan,
            ))
            nama_aset_list.append(aset.nama)

    catat_aktivitas(
        aksi="CREATE",
        target_model="Peminjaman",
        target_id=peminjaman.id,
        deskripsi=f"Peminjaman aset oleh {nama_peminjam}" + (f": {', '.join(nama_aset_list)}" if nama_aset_list else ""),
        data_baru={
            "nama_peminjam": nama_peminjam,
            "unit": unit,
            "lokasi_kerja": lokasi_kerja,
            "jenis_transaksi": jenis_transaksi,
            "barang": nama_aset_list,
            "tanggal_pinjam": tanggal_pinjam.strftime("%Y-%m-%d"),
            "tanggal_rencana_kembali": tanggal_rencana_kembali.strftime("%Y-%m-%d") if tanggal_rencana_kembali else None,
            "evidence": evidence,
        }
    )

    db.session.commit()
    flash(f"Peminjaman berhasil dibuat untuk {nama_peminjam}.", "success")
    return redirect(url_for("peminjaman_list"))


@app.route("/peminjaman/<int:id>/kembalikan", methods=["POST"])
@login_required
@role_required(ROLE_ADMIN)
def peminjaman_kembalikan(id):
    """Tandai peminjaman sebagai sudah dikembalikan."""
    peminjaman = Peminjaman.query.get_or_404(id)

    if peminjaman.status == "Dikembalikan":
        flash("Peminjaman ini sudah ditandai dikembalikan sebelumnya.", "warning")
        return redirect(url_for("peminjaman_list"))

    tanggal_kembali_str = request.form.get("tanggal_dikembalikan")
    if tanggal_kembali_str:
        try:
            peminjaman.tanggal_dikembalikan = datetime.strptime(tanggal_kembali_str, "%Y-%m-%d").date()
        except ValueError:
            peminjaman.tanggal_dikembalikan = datetime.now(WIB).date()
    else:
        peminjaman.tanggal_dikembalikan = datetime.now(WIB).date()

    peminjaman.status = "Dikembalikan"

    for pa in peminjaman.aset_terkait:
        aset = pa.aset
        if aset:
            db.session.add(HistoriAset(
                id_aset=aset.id,
                jenis_event="kembali",
                gedung=aset.gedung,
                lantai=aset.lantai,
                ruangan=aset.ruangan,
            ))

    catat_aktivitas(
        aksi="UPDATE",
        target_model="Peminjaman",
        target_id=peminjaman.id,
        deskripsi=f"Pengembalian aset dari peminjaman {peminjaman.nama_peminjam}",
        data_baru={"status": "Dikembalikan", "tanggal_dikembalikan": peminjaman.tanggal_dikembalikan.strftime("%Y-%m-%d")}
    )

    db.session.commit()
    flash("Aset berhasil ditandai sudah dikembalikan.", "success")
    return redirect(url_for("peminjaman_list"))


@app.route("/peminjaman/<int:id>/detail")
@login_required
def peminjaman_detail(id):
    """Detail satu data peminjaman aset."""
    peminjaman = Peminjaman.query.get_or_404(id)
    today = datetime.now(WIB).date()
    return render_template("peminjaman/detail.html", p=peminjaman, today=today)


@app.route("/peminjaman/<int:id>/evidence/upload", methods=["POST"])
@login_required
@role_required(ROLE_ADMIN)
def peminjaman_evidence_upload(id):
    """Upload evidence laporan (BA/PDF) tambahan untuk 1 peminjaman.
    Tidak menimpa evidence lama -- ditambahkan ke histori evidence_list."""
    peminjaman = Peminjaman.query.get_or_404(id)

    filename, error = save_evidence_pdf(request.files.get("evidence_baru"))
    if error:
        flash(f"Gagal upload evidence: {error}", "danger")
        return redirect(url_for("peminjaman_detail", id=id))

    ev = PeminjamanEvidence(
        id_peminjaman=peminjaman.id,
        filename=filename,
        keterangan=request.form.get("keterangan_evidence") or None,
        id_user_uploader=current_user.id,
    )
    db.session.add(ev)

    catat_aktivitas(
        aksi="UPDATE",
        target_model="Peminjaman",
        target_id=peminjaman.id,
        deskripsi=f"Upload evidence laporan baru untuk peminjaman oleh {peminjaman.nama_peminjam}",
        data_baru={"evidence": filename, "keterangan_evidence": ev.keterangan},
    )
    db.session.commit()
    flash("Evidence laporan berhasil diupload.", "success")
    return redirect(url_for("peminjaman_detail", id=id))


@app.route("/peminjaman/<int:id>/delete", methods=["POST"])
@login_required
@role_required(ROLE_ADMIN)
def peminjaman_delete(id):
    """Hapus data peminjaman aset."""
    peminjaman = Peminjaman.query.get_or_404(id)
    catat_aktivitas(
        aksi="DELETE",
        target_model="Peminjaman",
        target_id=peminjaman.id,
        deskripsi=f"Menghapus data peminjaman {peminjaman.nama_peminjam}",
        data_lama=snapshot_peminjaman(peminjaman),
    )
    db.session.delete(peminjaman)
    db.session.commit()
    flash("Data peminjaman berhasil dihapus.", "success")
    return redirect(url_for("peminjaman_list"))


def _redirect_aman(next_url, default_endpoint="dashboard"):
    """Redirect ke `next_url` kalau valid & lokal (mencegah open-redirect),
    kalau tidak fallback ke `default_endpoint`."""
    if next_url and next_url.startswith("/") and not next_url.startswith("//"):
        return redirect(next_url)
    return redirect(url_for(default_endpoint))


@app.route("/peminjaman/<int:id>/konfirmasi-perpanjangan", methods=["POST"])
@login_required
@role_required(ROLE_ADMIN)
def peminjaman_konfirmasi_perpanjangan(id):
    """Konfirmasi dari notifikasi H-10: apakah peminjaman diperpanjang atau tidak."""
    peminjaman = Peminjaman.query.get_or_404(id)
    keputusan = request.form.get("keputusan")  # 'perpanjang' atau 'tidak'
    next_url = request.form.get("next")  # kembali ke halaman asal (dashboard atau halaman reminder)

    if peminjaman.status != "Dipinjam":
        flash("Peminjaman ini sudah dikembalikan, tidak perlu konfirmasi perpanjangan.", "warning")
        return _redirect_aman(next_url)

    # --- Evidence laporan -----------------------------------------------
    # Evidence WAJIB diupload dan HARUS berupa PDF, berlaku untuk kedua
    # keputusan (diperpanjang maupun tidak diperpanjang) -- BA konfirmasi
    # perpanjangan harus selalu ada dokumen PDF resminya.
    evidence_filename, evidence_error = save_evidence_pdf(request.files.get("evidence_baru"))
    if evidence_error:
        flash(f"Gagal upload evidence: {evidence_error}", "danger")
        return _redirect_aman(next_url)

    if keputusan == "perpanjang":
        tanggal_baru_str = request.form.get("tanggal_baru")
        if not tanggal_baru_str:
            flash("Tanggal rencana kembali yang baru wajib diisi untuk perpanjangan.", "danger")
            return _redirect_aman(next_url)
        try:
            tanggal_baru = datetime.strptime(tanggal_baru_str, "%Y-%m-%d").date()
        except ValueError:
            flash("Format tanggal tidak valid.", "danger")
            return _redirect_aman(next_url)

        # VALIDASI URUTAN TANGGAL: tanggal perpanjangan baru tidak boleh
        # lebih awal dari tanggal pinjam awal.
        if peminjaman.tanggal_pinjam and tanggal_baru < peminjaman.tanggal_pinjam:
            flash("Tanggal rencana kembali baru tidak boleh lebih awal dari tanggal pinjam.", "danger")
            return _redirect_aman(next_url)

        tanggal_lama = peminjaman.tanggal_rencana_kembali
        peminjaman.tanggal_rencana_kembali = tanggal_baru
        # Reset supaya notifikasi bisa muncul lagi kalau tanggal baru juga mendekati H-10
        peminjaman.status_perpanjangan = None

        keterangan_evidence = (
            f"Evidence perpanjangan s.d. {tanggal_baru.strftime('%d-%m-%Y')}"
        )
        db.session.add(PeminjamanEvidence(
            id_peminjaman=peminjaman.id,
            filename=evidence_filename,
            keterangan=keterangan_evidence,
            id_user_uploader=current_user.id,
        ))

        data_baru = {"tanggal_rencana_kembali": tanggal_baru.strftime("%Y-%m-%d")}
        if evidence_filename:
            data_baru["evidence"] = evidence_filename
            data_baru["keterangan_evidence"] = keterangan_evidence

        catat_aktivitas(
            aksi="UPDATE",
            target_model="Peminjaman",
            target_id=peminjaman.id,
            deskripsi=(
                f"Perpanjangan peminjaman oleh {peminjaman.nama_peminjam}: "
                f"{tanggal_lama.strftime('%d-%m-%Y') if tanggal_lama else '-'} → {tanggal_baru.strftime('%d-%m-%Y')}"
            ),
            data_lama={"tanggal_rencana_kembali": tanggal_lama.strftime("%Y-%m-%d") if tanggal_lama else None},
            data_baru=data_baru,
        )
        db.session.commit()
        flash(
            f"Peminjaman oleh {peminjaman.nama_peminjam} berhasil diperpanjang sampai "
            f"{tanggal_baru.strftime('%d-%m-%Y')}.",
            "success",
        )

    elif keputusan == "tidak":
        peminjaman.status_perpanjangan = "Tidak Diperpanjang"

        keterangan_evidence = None
        if evidence_filename:
            keterangan_evidence = "Evidence konfirmasi tidak diperpanjang"
            db.session.add(PeminjamanEvidence(
                id_peminjaman=peminjaman.id,
                filename=evidence_filename,
                keterangan=keterangan_evidence,
                id_user_uploader=current_user.id,
            ))

        data_baru = {"status_perpanjangan": "Tidak Diperpanjang"}
        if evidence_filename:
            data_baru["evidence"] = evidence_filename
            data_baru["keterangan_evidence"] = keterangan_evidence

        catat_aktivitas(
            aksi="UPDATE",
            target_model="Peminjaman",
            target_id=peminjaman.id,
            deskripsi=(
                f"Konfirmasi TIDAK diperpanjang untuk peminjaman oleh {peminjaman.nama_peminjam} "
                f"— aset wajib dikembalikan sesuai jadwal."
            ),
            data_baru=data_baru,
        )
        db.session.commit()
        flash(f"Dikonfirmasi: peminjaman oleh {peminjaman.nama_peminjam} tidak diperpanjang.", "success")

    else:
        flash("Keputusan tidak valid.", "danger")

    return _redirect_aman(next_url)


@app.route("/peminjaman/reminder")
@login_required
@role_required(ROLE_ADMIN)
def peminjaman_reminder_page():
    """Halaman perbesar (fullscreen) dari widget notifikasi H-10 di dashboard.

    Sengaja TIDAK didaftarkan di sidebar -- hanya bisa diakses lewat tombol
    perbesar di widget dashboard. Menampilkan seluruh peminjaman yang rencana
    kembalinya sudah lewat (Terlambat) sampai yang tinggal 10 hari lagi,
    lengkap dengan filter & pagination.
    """
    today = datetime.now(WIB).date()
    batas_notif = today + timedelta(days=10)

    filter_waktu = request.args.get("filter", "").strip()  # '' / 'terlambat' / 'mendekati'
    search = request.args.get("search", "").strip()
    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 10, type=int)
    if per_page not in (10, 25, 50, 100):
        per_page = 10

    query = Peminjaman.query.filter(
        Peminjaman.status == "Dipinjam",
        Peminjaman.tanggal_rencana_kembali.isnot(None),
        Peminjaman.tanggal_rencana_kembali <= batas_notif,
        Peminjaman.status_perpanjangan.is_(None),
    )

    if filter_waktu == "terlambat":
        query = query.filter(Peminjaman.tanggal_rencana_kembali < today)
    elif filter_waktu == "mendekati":
        query = query.filter(Peminjaman.tanggal_rencana_kembali >= today)

    if search:
        query = query.outerjoin(PeminjamanAset).outerjoin(Aset, PeminjamanAset.id_aset == Aset.id).filter(
            db.or_(
                Peminjaman.nama_peminjam.ilike(f"%{search}%"),
                Peminjaman.unit.ilike(f"%{search}%"),
                Peminjaman.lokasi_kerja.ilike(f"%{search}%"),
                Aset.nama.ilike(f"%{search}%"),
            )
        ).distinct()

    query = query.order_by(Peminjaman.tanggal_rencana_kembali.asc())

    # Hitung total aset yang terlibat dari SELURUH hasil filter (bukan hanya 1 halaman)
    id_peminjaman_terfilter = [p.id for p in query.with_entities(Peminjaman.id).all()]
    total_aset_terfilter = 0
    if id_peminjaman_terfilter:
        total_aset_terfilter = (
            db.session.query(db.func.count(PeminjamanAset.id))
            .filter(PeminjamanAset.id_peminjaman.in_(id_peminjaman_terfilter))
            .scalar()
        ) or 0

    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    daftar_reminder = pagination.items

    total_terlambat = Peminjaman.query.filter(
        Peminjaman.status == "Dipinjam",
        Peminjaman.tanggal_rencana_kembali.isnot(None),
        Peminjaman.tanggal_rencana_kembali < today,
        Peminjaman.status_perpanjangan.is_(None),
    ).count()
    total_mendekati = Peminjaman.query.filter(
        Peminjaman.status == "Dipinjam",
        Peminjaman.tanggal_rencana_kembali.isnot(None),
        Peminjaman.tanggal_rencana_kembali >= today,
        Peminjaman.tanggal_rencana_kembali <= batas_notif,
        Peminjaman.status_perpanjangan.is_(None),
    ).count()

    return render_template(
        "peminjaman/reminder.html",
        daftar_reminder=daftar_reminder,
        pagination=pagination,
        filter_waktu=filter_waktu,
        search=search,
        today=today,
        total_terlambat=total_terlambat,
        total_mendekati=total_mendekati,
        total_semua=total_terlambat + total_mendekati,
        total_aset_terfilter=total_aset_terfilter,
    )


@app.route("/peminjaman/import-guide")
@login_required
@role_required(ROLE_ADMIN)
def peminjaman_import_guide():
    """Halaman panduan + form upload import Excel Peminjaman."""
    return render_template("peminjaman/import_guide.html")


@app.route("/peminjaman/import", methods=["POST"])
@login_required
@role_required(ROLE_ADMIN)
def peminjaman_import():
    """Import data Peminjaman dari file Excel (gaya sheet 'BA Transfer').

    Sheet & baris header dideteksi otomatis berdasarkan nama kolom (Nama,
    Unit, Lokasi Kerja, Jenis Barang, Jenis Transaksi, Tanggal Awal,
    Tanggal Akhir, Keterangan, Evidence Lampiran) -- bukan posisi tetap --
    supaya file yang punya banyak sheet lain (mis. data Aset) tetap bisa
    dipakai, sheet Peminjaman-nya otomatis ketemu.

    Catatan desain (hasil konfirmasi dengan user):
    - Jenis Barang disimpan sebagai teks bebas, TIDAK ditambahkan ke tabel
      Aset utama.
    - Evidence Lampiran disimpan sebagai link Google Drive (evidence_link),
      bukan file yang didownload ke server.
    - Unit disimpan sebagai teks bebas di kolom Peminjaman.unit, TIDAK ada
      tabel master Unit terpisah.
    """
    file = request.files.get("file_import")
    if not file or file.filename == "":
        flash("Pilih file Excel (.xlsx) terlebih dahulu.", "danger")
        return redirect(url_for("peminjaman_import_guide"))
    if not file.filename.lower().endswith((".xlsx", ".xlsm")):
        flash("Format file harus .xlsx.", "danger")
        return redirect(url_for("peminjaman_import_guide"))

    try:
        wb = openpyxl.load_workbook(file.stream, data_only=True)
    except Exception:
        flash("File Excel tidak valid atau rusak.", "danger")
        return redirect(url_for("peminjaman_import_guide"))

    ws, header_row, col_map = find_peminjaman_sheet(wb)
    if not ws:
        flash(
            "Tidak ditemukan sheet dengan format kolom Peminjaman (Nama, "
            "Jenis Transaksi, Evidence Lampiran, dst) di file ini. Pastikan "
            "ada sheet dengan header kolom yang sesuai (lihat panduan).",
            "danger",
        )
        return redirect(url_for("peminjaman_import_guide"))

    def col(field, row):
        idx = col_map.get(field)
        return ws.cell(row=row, column=idx) if idx else None

    ditambahkan, dilewati, tanpa_evidence_link = 0, 0, 0
    error_baris = []

    for row_idx in range(header_row + 1, (ws.max_row or header_row) + 1):
        cells_in_row = [ws.cell(row=row_idx, column=idx) for idx in col_map.values()]
        if not any(c.value not in (None, "") for c in cells_in_row):
            continue  # baris kosong total, lewati diam-diam

        nama_cell = col("nama", row_idx)
        unit_cell = col("unit", row_idx)
        lokasi_cell = col("lokasi_kerja", row_idx)
        jenis_barang_cell = col("jenis_barang", row_idx)
        jenis_transaksi_cell = col("jenis_transaksi", row_idx)
        tanggal_awal_cell = col("tanggal_awal", row_idx)
        tanggal_akhir_cell = col("tanggal_akhir", row_idx)
        keterangan_cell = col("keterangan", row_idx)
        evidence_cell = col("evidence_lampiran", row_idx)

        def txt(cell):
            if cell is None or cell.value in (None, ""):
                return ""
            return str(cell.value).strip()

        nama = txt(nama_cell) or "-"  # beberapa baris (mis. Pelimpahan IN) tidak punya nama
        unit_nama = txt(unit_cell)
        lokasi_kerja = txt(lokasi_cell)
        jenis_barang = txt(jenis_barang_cell)
        jenis_transaksi = txt(jenis_transaksi_cell)
        keterangan = txt(keterangan_cell)

        tanggal_awal = parse_excel_date(tanggal_awal_cell.value if tanggal_awal_cell else None)
        tanggal_akhir = parse_excel_date(tanggal_akhir_cell.value if tanggal_akhir_cell else None)

        # tanggal_pinjam wajib diisi di database. Kalau Tanggal Awal kosong
        # (umum terjadi pada baris "Pengembalian" di data sumber), pakai
        # Tanggal Akhir sebagai fallback.
        tanggal_pinjam = tanggal_awal or tanggal_akhir
        if not tanggal_pinjam:
            dilewati += 1
            error_baris.append(f"Baris {row_idx}: Tanggal Awal & Tanggal Akhir kosong, dilewati")
            continue

        is_pengembalian = "pengembalian" in jenis_transaksi.lower()
        if is_pengembalian:
            status = "Dikembalikan"
            tanggal_rencana_kembali = None
            tanggal_dikembalikan = tanggal_akhir or tanggal_awal
        else:
            status = "Dipinjam"
            tanggal_rencana_kembali = tanggal_akhir if tanggal_awal else None
            tanggal_dikembalikan = None

        evidence_link = extract_link_from_cell(evidence_cell)
        if not evidence_link:
            tanpa_evidence_link += 1

        # Catatan: Unit TIDAK disimpan sebagai tabel master terpisah -- cukup
        # teks bebas di kolom Peminjaman.unit (nilai untuk dropdown filter
        # Unit di halaman Peminjaman diambil otomatis dari nilai-nilai unik
        # kolom ini, bukan dari tabel master).

        # Jenis Barang dari Excel HANYA ditautkan ke Kategori yang SUDAH ADA
        # (terdaftar lewat modul Data Aset) -- TIDAK PERNAH membuat Kategori
        # baru dari sini, supaya daftar Kategori tidak ikut membengkak tiap
        # kali import Excel Peminjaman. Teks aslinya tetap disimpan di
        # jenis_barang (legacy) walaupun tidak ketemu kategorinya.
        kategori_obj = find_kategori_saja(jenis_barang) if jenis_barang else None

        db.session.add(Peminjaman(
            nama_peminjam=nama,
            unit=unit_nama or None,
            lokasi_kerja=lokasi_kerja or None,
            jenis_barang=jenis_barang or None,
            id_kategori=kategori_obj.id if kategori_obj else None,
            jenis_transaksi=jenis_transaksi or None,
            evidence_link=evidence_link,
            sumber_import=True,
            tanggal_pinjam=tanggal_pinjam,
            tanggal_rencana_kembali=tanggal_rencana_kembali,
            tanggal_dikembalikan=tanggal_dikembalikan,
            status=status,
            keterangan=keterangan or None,
            created_by=current_user.id,
        ))
        ditambahkan += 1

    if ditambahkan:
        catat_aktivitas(
            aksi="IMPORT",
            target_model="Peminjaman",
            target_id=0,
            deskripsi=f"Import Excel peminjaman dari sheet '{ws.title}': {ditambahkan} baris ditambahkan",
        )

    db.session.commit()

    pesan = f"Import selesai dari sheet '{ws.title}': {ditambahkan} data peminjaman ditambahkan."
    if dilewati:
        pesan += f" {dilewati} baris dilewati (tanggal kosong)."
    flash(pesan, "warning" if dilewati else "success")
    if tanpa_evidence_link:
        flash(
            f"Perhatian: {tanpa_evidence_link} baris tidak punya link evidence yang "
            f"terdeteksi (cell tidak ada hyperlink Google Drive maupun teks link).",
            "warning",
        )
    if error_baris:
        flash(" | ".join(error_baris[:5]), "warning")

    return redirect(url_for("peminjaman_list"))


@app.route("/maintenance")
@login_required
@role_required(ROLE_ADMIN)
def maintenance_list():
    kategori_id = request.args.get("kategori", "").strip()
    search = request.args.get("search", "").strip()
    status = request.args.get("status", "").strip()
    
    query = Maintenance.query.join(Aset)
    
    if kategori_id:
        # PERBAIKAN: filter berdasarkan kategori ASET SAAT INI (Aset.id_kategori),
        # bukan Maintenance.kategori (teks snapshot lama yang sudah tidak sinkron
        # sejak kategori dipecah/diubah -- mis. "Furniture" -> "Office Furniture").
        query = query.filter(Aset.id_kategori == kategori_id)
    if status:
        query = query.filter(Maintenance.status == status)
    if search:
        query = query.filter(
            db.or_(
                Aset.nama.ilike(f"%{search}%"),
                Aset.kode_aset.ilike(f"%{search}%"),
                Maintenance.judul.ilike(f"%{search}%")
            )
        )
    
    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 10, type=int)
    pagination = query.order_by(Maintenance.tanggal_mulai.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    daftar_maintenance = pagination.items
    aset_all = Aset.query.order_by(Aset.nama).all()
    
    # +++ TAMBAHAN: Ambil semua kategori aset untuk filter +++
    kategori_all = Kategori.query.order_by(Kategori.nama).all()

    gedung_all = (
        db.session.query(Aset.area, Aset.gedung)
        .filter(Aset.gedung.isnot(None), Aset.gedung != "")
        .distinct()
        .order_by(Aset.area, Aset.gedung)
        .all()
    )
    gedung_all_formatted = [
        {"value": g.gedung, "label": f"{format_area_label(g.area)} - {g.gedung}" if g.area else g.gedung}
        for g in gedung_all
    ]
    
    return render_template(
        "maintenance/list.html",
        daftar_maintenance=daftar_maintenance,
        pagination=pagination,
        aset_all=aset_all,
        kategori_all=kategori_all,  # <-- KIRIM KE TEMPLATE
        kategori_terpilih=kategori_id,
        status_terpilih=status,
        search=search,
        gedung_all=gedung_all_formatted,
    )


@app.route("/maintenance/export")
@login_required
@role_required(ROLE_ADMIN)
def maintenance_export():
    """Export seluruh data Maintenance ke Excel. Kolom Kategori Aset diambil
    dari kategori ASET SAAT INI (Aset.kategori_ref), bukan Maintenance.kategori
    (teks snapshot lama yang bisa sudah tidak sinkron -- sama seperti alasan
    perbaikan filter kategori di maintenance_list)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data Maintenance"
    ws.append([
        "No", "Kode Aset", "Nama Aset", "Kategori Aset",
        "Judul Maintenance", "Tipe", "Vendor",
        "Tanggal Mulai", "Tanggal Akhir", "Biaya (Rp)", "Status",
        "Deskripsi", "Dibuat Oleh", "Tanggal Dibuat",
    ])

    daftar = Maintenance.query.join(Aset).order_by(Maintenance.tanggal_mulai.desc()).all()
    for no, m in enumerate(daftar, start=1):
        ws.append([
            no,
            m.aset.kode_aset if m.aset else "",
            m.aset.nama if m.aset else "",
            m.aset.kategori_ref.nama if m.aset and m.aset.kategori_ref else "",
            m.judul,
            m.tipe,
            m.vendor or "",
            m.tanggal_mulai.strftime("%Y-%m-%d") if m.tanggal_mulai else "",
            m.tanggal_akhir.strftime("%Y-%m-%d") if m.tanggal_akhir else "",
            float(m.biaya) if m.biaya else 0,
            m.status,
            m.deskripsi or "",
            m.user.name if m.user else "-",
            m.created_at.strftime("%Y-%m-%d %H:%M") if m.created_at else "",
        ])

    for i in range(1, 15):
        ws.column_dimensions[get_column_letter(i)].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    nama_file = f"data_maintenance_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=nama_file,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/maintenance/create", methods=["POST"])
@login_required
@role_required(ROLE_ADMIN)
def maintenance_create():
    """Tambah jadwal maintenance baru."""
    # Ambil dari radio button (single aset)
    aset_id = request.form.get("aset_ids")
    if not aset_id:
        flash("Pilih aset terlebih dahulu.", "danger")
        return redirect(url_for("maintenance_list"))
    
    # Ambil kategori dari form (sudah otomatis terisi dari JavaScript)
    kategori = request.form.get("kategori", "").strip()
    judul = request.form.get("judul", "").strip()
    deskripsi = request.form.get("deskripsi", "").strip()
    vendor = request.form.get("vendor", "").strip()
    tipe = request.form.get("tipe", "Preventif")
    tanggal_mulai_str = request.form.get("tanggal_mulai")
    tanggal_akhir_str = request.form.get("tanggal_akhir")
    biaya = request.form.get("biaya", 0)
    status = request.form.get("status", "Scheduled")
    
    # Validasi
    if not judul or not tanggal_mulai_str:
        flash("Judul dan tanggal mulai wajib diisi.", "danger")
        return redirect(url_for("maintenance_list"))
    
    # Validasi aset
    aset = Aset.query.get(aset_id)
    if not aset:
        flash("Aset tidak ditemukan.", "danger")
        return redirect(url_for("maintenance_list"))
    
    # Jika kategori dari form kosong, ambil dari kategori aset (fallback)
    if not kategori:
        kategori = aset.kategori_ref.nama if aset.kategori_ref else "Tidak ada kategori"
    
    # Parse tanggal
    try:
        tanggal_mulai = datetime.strptime(tanggal_mulai_str, "%Y-%m-%d").date()
    except ValueError:
        flash("Format tanggal mulai tidak valid.", "danger")
        return redirect(url_for("maintenance_list"))
    
    tanggal_akhir = None
    if tanggal_akhir_str:
        try:
            tanggal_akhir = datetime.strptime(tanggal_akhir_str, "%Y-%m-%d").date()
        except ValueError:
            pass
    
    # Buat maintenance
    maintenance = Maintenance(
        id_aset=int(aset_id),
        kategori=kategori,
        judul=judul,
        deskripsi=deskripsi or None,
        vendor=vendor or None,
        tipe=tipe,
        tanggal_mulai=tanggal_mulai,
        tanggal_akhir=tanggal_akhir,
        biaya=float(biaya) if biaya else 0,
        status=status,
        created_by=current_user.id,
    )
    db.session.add(maintenance)
    db.session.flush()  # penting: dapatkan maintenance.id dari MySQL (AUTO_INCREMENT) sebelum dipakai di log aktivitas
    
    # Catat histori aset
    histori = HistoriAset(
        id_aset=int(aset_id),
        jenis_event="maintenance",
        gedung=aset.gedung,
        lantai=aset.lantai,
        ruangan=aset.ruangan,
        id_tiket=None
    )
    db.session.add(histori)
    
    # Catat aktivitas admin
    catat_aktivitas(
        aksi="CREATE",
        target_model="Maintenance",
        target_id=maintenance.id,
        deskripsi=f"Menambahkan jadwal maintenance untuk aset {aset.nama} ({aset.kode_aset}): {judul}",
        data_baru=snapshot_maintenance(maintenance),
    )
    
    db.session.commit()
    flash(f"Jadwal maintenance berhasil ditambahkan untuk aset {aset.nama}.", "success")
    return redirect(url_for("maintenance_list"))

@app.route("/maintenance/<int:id>/edit", methods=["POST"])
@login_required
@role_required(ROLE_ADMIN)
def maintenance_edit(id):
    """Edit jadwal maintenance."""
    maintenance = Maintenance.query.get_or_404(id)
    data_lama = snapshot_maintenance(maintenance)

    maintenance.judul = request.form.get("judul", "").strip()
    maintenance.deskripsi = request.form.get("deskripsi", "").strip() or None
    maintenance.vendor = request.form.get("vendor", "").strip() or None
    maintenance.tipe = request.form.get("tipe", "Preventif")
    biaya_raw = request.form.get("biaya", "").strip()
    maintenance.biaya = float(biaya_raw) if biaya_raw else 0
    maintenance.status = request.form.get("status", "Scheduled")

    kategori_raw = request.form.get("kategori", "").strip()
    if kategori_raw:
        kategori_map = {"elektronik": "Elektronik", "furniture": "Furniture", "lainnya": "Lainnya"}
        maintenance.kategori = kategori_map.get(kategori_raw.lower(), kategori_raw.capitalize())
    
    tanggal_akhir_str = request.form.get("tanggal_akhir")
    if tanggal_akhir_str:
        try:
            maintenance.tanggal_akhir = datetime.strptime(tanggal_akhir_str, "%Y-%m-%d").date()
        except ValueError:
            pass

    data_baru = snapshot_maintenance(maintenance)
    if data_lama != data_baru:
        catat_aktivitas(
            aksi="UPDATE",
            target_model="Maintenance",
            target_id=maintenance.id,
            deskripsi=f"Memperbarui jadwal maintenance: {maintenance.judul}",
            data_lama=data_lama,
            data_baru=data_baru,
        )

    db.session.commit()
    flash("Jadwal maintenance berhasil diperbarui.", "success")
    return redirect(url_for("maintenance_list"))


@app.route("/maintenance/<int:id>/detail")
@login_required
def maintenance_detail(id):
    """Halaman detail sebuah jadwal maintenance, termasuk foto dokumentasi
    (before / on progress / after). Foto dokumentasi HANYA tampil & bisa
    diupload dari halaman ini, tidak ada di halaman list."""
    maintenance = Maintenance.query.get_or_404(id)
    return render_template("maintenance/detail.html", m=maintenance)


# Slot foto dokumentasi yang diizinkan + nama kolom di model Maintenance
MAINTENANCE_FOTO_SLOTS = {
    "before": "foto_before",
    "progress": "foto_progress",
    "after": "foto_after",
}


@app.route("/maintenance/<int:id>/foto/<slot>", methods=["POST"])
@login_required
@role_required(ROLE_ADMIN)
def maintenance_upload_foto(id, slot):
    """Upload/ganti satu foto dokumentasi (before/progress/after) milik
    sebuah jadwal maintenance. Selalu redirect balik ke halaman detail."""
    maintenance = Maintenance.query.get_or_404(id)

    if slot not in MAINTENANCE_FOTO_SLOTS:
        flash("Jenis foto tidak dikenali.", "danger")
        return redirect(url_for("maintenance_detail", id=id))

    kolom = MAINTENANCE_FOTO_SLOTS[slot]
    file_storage = request.files.get("foto")

    if not file_storage or file_storage.filename == "":
        flash("Pilih file foto terlebih dahulu.", "danger")
        return redirect(url_for("maintenance_detail", id=id))

    unique_name, error = save_upload(file_storage, prefix=f"maintenance_{slot}_")
    if error:
        flash(error, "danger")
        return redirect(url_for("maintenance_detail", id=id))

    # Hapus file foto lama (kalau ada) supaya tidak menumpuk di server
    foto_lama = getattr(maintenance, kolom)
    if foto_lama:
        old_path = os.path.join(app.config["UPLOAD_FOLDER"], foto_lama)
        if os.path.exists(old_path):
            try:
                os.remove(old_path)
            except OSError:
                pass

    setattr(maintenance, kolom, unique_name)

    label_slot = {"before": "Sebelum", "progress": "Sedang Berlangsung", "after": "Sesudah"}[slot]
    catat_aktivitas(
        aksi="UPDATE",
        target_model="Maintenance",
        target_id=maintenance.id,
        deskripsi=f"Mengunggah foto dokumentasi ({label_slot}) untuk maintenance: {maintenance.judul}",
        data_lama={"foto": foto_lama, "label": label_slot} if foto_lama else None,
        data_baru={"foto": unique_name, "label": label_slot},
    )

    db.session.commit()
    flash(f"Foto dokumentasi ({label_slot}) berhasil disimpan.", "success")
    return redirect(url_for("maintenance_detail", id=id))


@app.route("/maintenance/<int:id>/foto/<slot>/delete", methods=["POST"])
@login_required
@role_required(ROLE_ADMIN)
def maintenance_delete_foto(id, slot):
    """Hapus satu foto dokumentasi (before/progress/after)."""
    maintenance = Maintenance.query.get_or_404(id)

    if slot not in MAINTENANCE_FOTO_SLOTS:
        flash("Jenis foto tidak dikenali.", "danger")
        return redirect(url_for("maintenance_detail", id=id))

    kolom = MAINTENANCE_FOTO_SLOTS[slot]
    foto_lama = getattr(maintenance, kolom)
    if foto_lama:
        old_path = os.path.join(app.config["UPLOAD_FOLDER"], foto_lama)
        if os.path.exists(old_path):
            try:
                os.remove(old_path)
            except OSError:
                pass
        setattr(maintenance, kolom, None)

        label_slot = {"before": "Sebelum", "progress": "Sedang Berlangsung", "after": "Sesudah"}[slot]
        catat_aktivitas(
            aksi="UPDATE",
            target_model="Maintenance",
            target_id=maintenance.id,
            deskripsi=f"Menghapus foto dokumentasi ({label_slot}) untuk maintenance: {maintenance.judul}",
            data_lama={"foto": foto_lama, "label": label_slot},
        )
        db.session.commit()
        flash("Foto dokumentasi berhasil dihapus.", "success")

    return redirect(url_for("maintenance_detail", id=id))


@app.route("/maintenance/<int:id>/delete", methods=["POST"])
@login_required
@role_required(ROLE_ADMIN)
def maintenance_delete(id):
    """Hapus jadwal maintenance."""
    maintenance = Maintenance.query.get_or_404(id)
    judul = maintenance.judul
    nama_aset = maintenance.aset.nama if maintenance.aset else "-"
    data_lama = snapshot_maintenance(maintenance)

    # Bersihkan file foto dokumentasi dari disk juga
    for kolom in MAINTENANCE_FOTO_SLOTS.values():
        foto = getattr(maintenance, kolom)
        if foto:
            path = os.path.join(app.config["UPLOAD_FOLDER"], foto)
            if os.path.exists(path):
                try:
                    os.remove(path)
                except OSError:
                    pass

    catat_aktivitas(
        aksi="DELETE",
        target_model="Maintenance",
        target_id=maintenance.id,
        deskripsi=f"Menghapus jadwal maintenance: {judul} (aset {nama_aset})",
        data_lama=data_lama,
    )

    db.session.delete(maintenance)
    db.session.commit()
    flash("Jadwal maintenance berhasil dihapus.", "success")
    return redirect(url_for("maintenance_list"))

# ---------------------------------------------------------------------------
# USERS (Admin only)
# ---------------------------------------------------------------------------
@app.route("/users")
@login_required
@role_required(ROLE_ADMIN)
def users_list():
    daftar_user = User.query.all()
    return render_template(
        "users/list.html",
        daftar_user=daftar_user,
        now=datetime.now(WIB).replace(tzinfo=None),
    )


@app.route("/users/create", methods=["POST"])
@login_required
@role_required(ROLE_ADMIN)
def users_create():
    email = request.form.get("email", "").strip()
    password = request.form.get("password", "")
    role = request.form.get("role")

    if User.query.filter_by(email=email).first():
        flash("Email sudah terdaftar.", "danger")
        return redirect(url_for("users_list"))
    if len(password) < 8:
        flash("Password minimal 8 karakter.", "danger")
        return redirect(url_for("users_list"))
    
    if role not in ["admin", "user"]:
        flash("Role tidak valid.", "danger")
        return redirect(url_for("users_list"))

    db.session.add(User(
        name=request.form.get("name"),
        email=email,
        password=generate_password_hash(password),
        role=role,
    ))
    db.session.commit()
    flash("User berhasil ditambahkan.", "success")
    return redirect(url_for("users_list"))


@app.route("/users/<int:user_id>/toggle", methods=["POST"])
@login_required
@role_required(ROLE_ADMIN)
def users_toggle(user_id):
    if user_id == current_user.id:
        flash("Tidak bisa menonaktifkan akun sendiri.", "danger")
        return redirect(url_for("users_list"))
    user = User.query.get_or_404(user_id)
    user.is_active = not user.is_active
    db.session.commit()
    flash(
        f"User {'diaktifkan kembali' if user.is_active else 'dinonaktifkan'}.",
        "success",
    )
    return redirect(url_for("users_list"))


@app.route("/users/<int:user_id>/ban", methods=["POST"])
@login_required
@role_required(ROLE_ADMIN)
def users_ban(user_id):
    if user_id == current_user.id:
        flash("Tidak bisa mem-ban akun sendiri.", "danger")
        return redirect(url_for("users_list"))

    user = User.query.get_or_404(user_id)
    banned_until_str = request.form.get("banned_until", "").strip()
    ban_reason = request.form.get("ban_reason", "").strip() or None

    if not banned_until_str:
        flash("Tentukan tanggal & waktu berakhirnya ban.", "danger")
        return redirect(url_for("users_list"))
    try:
        banned_until = datetime.strptime(banned_until_str, "%Y-%m-%dT%H:%M")
    except ValueError:
        flash("Format tanggal/waktu ban tidak valid.", "danger")
        return redirect(url_for("users_list"))

    if banned_until <= datetime.now(WIB).replace(tzinfo=None):
        flash("Waktu berakhirnya ban harus di masa depan.", "danger")
        return redirect(url_for("users_list"))

    user.banned_until = banned_until
    user.ban_reason = ban_reason
    db.session.commit()
    flash(
        f"{user.name} di-ban sementara sampai {banned_until.strftime('%d-%m-%Y %H:%M')}.",
        "success",
    )
    return redirect(url_for("users_list"))


@app.route("/users/<int:user_id>/unban", methods=["POST"])
@login_required
@role_required(ROLE_ADMIN)
def users_unban(user_id):
    user = User.query.get_or_404(user_id)
    user.banned_until = None
    user.ban_reason = None
    db.session.commit()
    flash(f"Ban untuk {user.name} sudah dicabut.", "success")
    return redirect(url_for("users_list"))


# ---------------------------------------------------------------------------
# PROFIL (akun sendiri, semua role bisa akses)
# ---------------------------------------------------------------------------
@app.route("/profile")
@login_required
def profile():
    return render_template("profile.html")


@app.route("/profile/update", methods=["POST"])
@login_required
def profile_update():
    name = request.form.get("name", "").strip()
    email = request.form.get("email", "").strip()

    if not name:
        flash("Nama tidak boleh kosong.", "danger")
        return redirect(url_for("profile"))
    if not email:
        flash("Email tidak boleh kosong.", "danger")
        return redirect(url_for("profile"))

    # Cek email tidak dipakai user lain
    existing = User.query.filter(
        db.func.lower(User.email) == email.lower(), User.id != current_user.id
    ).first()
    if existing:
        flash("Email sudah dipakai akun lain.", "danger")
        return redirect(url_for("profile"))

    current_user.name = name
    current_user.email = email
    db.session.commit()
    flash("Profil berhasil diperbarui.", "success")
    return redirect(url_for("profile"))


@app.route("/profile/password", methods=["POST"])
@login_required
def profile_password():
    password_lama = request.form.get("password_lama", "")
    password_baru = request.form.get("password_baru", "")
    password_konfirmasi = request.form.get("password_konfirmasi", "")

    if not check_password_hash(current_user.password, password_lama):
        flash("Password lama tidak sesuai.", "danger")
        return redirect(url_for("profile"))
    if len(password_baru) < 8:
        flash("Password baru minimal 8 karakter.", "danger")
        return redirect(url_for("profile"))
    if password_baru != password_konfirmasi:
        flash("Konfirmasi password baru tidak cocok.", "danger")
        return redirect(url_for("profile"))

    current_user.password = generate_password_hash(password_baru)
    db.session.commit()
    flash("Password berhasil diubah.", "success")
    return redirect(url_for("profile"))


@app.route("/profile/photo", methods=["POST"])
@login_required
def profile_photo():
    # Hapus foto profil
    if request.form.get("hapus_foto"):
        if current_user.foto_profil:
            old_path = os.path.join(app.config["UPLOAD_FOLDER"], current_user.foto_profil)
            if os.path.exists(old_path):
                try:
                    os.remove(old_path)
                except Exception:
                    pass
            current_user.foto_profil = None
            db.session.commit()
            flash("Foto profil dihapus.", "success")
        return redirect(url_for("profile"))

    # Upload foto profil baru
    foto_file = request.files.get("foto_profil")
    if not foto_file or foto_file.filename == "":
        flash("Pilih file foto terlebih dahulu.", "danger")
        return redirect(url_for("profile"))

    unique_name, error = save_upload(foto_file, prefix=f"profil_{current_user.id}_")
    if error:
        flash(error, "danger")
        return redirect(url_for("profile"))

    # Hapus file foto lama supaya tidak menumpuk file yatim di server
    if current_user.foto_profil:
        old_path = os.path.join(app.config["UPLOAD_FOLDER"], current_user.foto_profil)
        if os.path.exists(old_path):
            try:
                os.remove(old_path)
            except Exception:
                pass

    current_user.foto_profil = unique_name
    db.session.commit()
    flash("Foto profil berhasil diperbarui.", "success")
    return redirect(url_for("profile"))


if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5001)